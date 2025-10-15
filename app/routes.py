from flask import Blueprint, render_template, request, redirect, url_for, session, flash, abort, Response, stream_with_context, jsonify, current_app
from .models import (
    Utilisateur,
    Evenement,
    FicheImplique,
    Bagage,
    ShareLink,
    ShareLinkAccessLog,
    Ticket,
    Animal,
    AuditLog,
    TimelineEntry,
    utilisateur_evenement,
    EventNews,
    BroadcastNotification,
)
from .extensions import db, limiter
from werkzeug.security import check_password_hash
from functools import wraps
from .audit import log_action
from datetime import datetime, timedelta, date, timezone
from collections import Counter
from flask_login import current_user
from flask import make_response
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
import io
from .backup_utils import is_db_empty, backup_to_bytesio, wipe_db, bulk_restore
from flask import send_file
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.units import cm, mm
import os
from io import BytesIO
import re
import json
import tempfile
import redis
from sqlalchemy import text, func, or_
import typing
import unicodedata
import threading

main_bp = Blueprint("main_bp", __name__)


BROADCAST_AUTO_CLEAR_SECONDS = 15
BROADCAST_ALLOWED_EMOJIS = ["ℹ️", "⚠️", "🚨", "✅", "🔥"]
BROADCAST_ALLOWED_LEVELS = ["info", "warning", "danger", "success", "critical"]
BROADCAST_LEVEL_LABELS = {
    "info": "Information",
    "warning": "Alerte",
    "danger": "Urgence",
    "success": "Succès",
    "critical": "Critique",
}
BROADCAST_DEFAULT_EMOJI = "⚠️"
BROADCAST_DEFAULT_LEVEL = "warning"


def schedule_broadcast_expiration(notification_id: int, delay: int = BROADCAST_AUTO_CLEAR_SECONDS) -> None:
    app = current_app._get_current_object()

    def _clear() -> None:
        with app.app_context():
            try:
                updated = (
                    BroadcastNotification.query.filter_by(id=notification_id, is_active=True)
                    .update({"is_active": False})
                )
                if updated:
                    db.session.commit()
                    app.logger.info(
                        "Broadcast notification %s automatically cleared after %s seconds",
                        notification_id,
                        delay,
                    )
                else:
                    db.session.rollback()
            except Exception:
                db.session.rollback()
                app.logger.exception(
                    "Unable to auto-clear broadcast notification %s",
                    notification_id,
                )

    timer = threading.Timer(delay, _clear)
    timer.daemon = True
    timer.start()


COMPETENCES_CAI: list[str] = [
    "Médecin",
    "Infirmier",
    "Sapeur-pompier",
    "SST",
    "Psychologue",
    "Bénévole",
    "Artisan",
    "Interprète",
    "Logisticien",
    "Conducteur",
    "Agent sécurité",
    "Autre",
]


def _slugify_competence(value: str) -> str:
    base = unicodedata.normalize("NFKD", value or "")
    base = "".join(ch for ch in base if not unicodedata.combining(ch))
    base = base.lower()
    return re.sub(r"[^a-z0-9]+", "", base)


_KNOWN_COMPETENCE_SLUGS: dict[str, str] = {
    _slugify_competence(name): name for name in COMPETENCES_CAI if name
}


def _canonicalize_competence_label(label: str) -> str | None:
    """Nettoie un libellé brut et le rapproche d'un intitulé connu."""

    if not label:
        return None

    cleaned = unicodedata.normalize("NFKC", label).replace("\r", "\n")
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" -–—•·:\t")
    cleaned = cleaned.strip()
    if not cleaned:
        return None

    slug = _slugify_competence(cleaned)
    if not slug:
        return None

    if slug in _KNOWN_COMPETENCE_SLUGS:
        return _KNOWN_COMPETENCE_SLUGS[slug]

    for known_slug, display in _KNOWN_COMPETENCE_SLUGS.items():
        if len(slug) >= 2 and known_slug.startswith(slug):
            return display
        if len(slug) >= 3 and slug.startswith(known_slug):
            return display
        if len(slug) >= 3 and slug in known_slug:
            return display

    if len(slug) <= 1:
        return None

    if len(cleaned) > 60:
        cleaned = cleaned[:57].rstrip() + "…"
    return cleaned

def add_timeline(fiche_id: int, user_id: int | None, content: str, kind: str):
    """Ajoute une entrée de timeline (UTC) et ne commit PAS (laisse l'appelant décider)."""
    e = TimelineEntry(
        fiche_id=fiche_id,
        user_id=user_id,
        content=(content or "").strip(),
        kind=(kind or "comment"),
        created_at=datetime.utcnow(),
    )
    db.session.add(e)
    return e




def json_nocache(payload: dict, status: int = 200):
    """Réponse JSON avec no-store pour empêcher tout cache navigateur/proxy."""
    resp = make_response(jsonify(payload), status)
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp

# 🔒 Décorateur pour vérifier l’authentification
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("main_bp.login"))
        return f(*args, **kwargs)
    return decorated_function

# 🔧 Fonction utilitaire
def get_current_user():
    return Utilisateur.query.get(session["user_id"])


AUDIT_LOG_RETENTION = timedelta(days=40)
PROVISIONAL_ACCOUNT_LIFETIME = timedelta(days=5)
_HOUSEKEEPING_STATE: dict[str, datetime | None] = {"last_run": None}


def _purge_old_audit_logs() -> int:
    cutoff = datetime.utcnow() - AUDIT_LOG_RETENTION
    deleted = AuditLog.query.filter(AuditLog.created_at < cutoff).delete(synchronize_session=False)
    if deleted:
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
            deleted = 0
    return deleted


def _ensure_provisional_deadlines() -> int:
    # Ensure every provisional account has an expiration
    updated = 0
    query = Utilisateur.query.filter(
        func.lower(Utilisateur.type_utilisateur) == "provisoire",
        Utilisateur.provisional_expires_at.is_(None),
    )
    now = datetime.utcnow()
    for account in query:
        base = account.created_at or now
        account.provisional_expires_at = base + PROVISIONAL_ACCOUNT_LIFETIME
        updated += 1
    if updated:
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
            updated = 0
    return updated


def _purge_expired_provisional_accounts() -> int:
    now = datetime.utcnow()
    expired = Utilisateur.query.filter(
        func.lower(Utilisateur.type_utilisateur) == "provisoire",
        Utilisateur.provisional_expires_at.isnot(None),
        Utilisateur.provisional_expires_at <= now,
    ).all()

    if not expired:
        return 0

    removed = 0
    for account in expired:
        removed += 1
        db.session.delete(account)

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return 0

    return removed


@main_bp.before_app_request
def _run_housekeeping():
    now = datetime.utcnow()
    last_run = _HOUSEKEEPING_STATE.get("last_run")
    if last_run and (now - last_run) < timedelta(hours=1):
        return

    try:
        _ensure_provisional_deadlines()
        _purge_expired_provisional_accounts()
        _purge_old_audit_logs()
    finally:
        _HOUSEKEEPING_STATE["last_run"] = now


# 🔐 Page de connexion
LOGIN_LOCK_THRESHOLD = 3
LOGIN_LOCK_DURATION = timedelta(minutes=5)
# Les tentatives sont suivies par adresse IP pour éviter de cibler un compte spécifique.
ip_login_attempts: dict[str, dict[str, datetime | int]] = {}

TIMELINE_COMMENT_MAX_LENGTH = 1000


def _get_client_ip() -> str:
    """Retourne l'adresse IP du client en tenant compte de X-Forwarded-For."""
    forwarded_for = request.headers.get("X-Forwarded-For", "").split(",")[0].strip()
    if forwarded_for:
        return forwarded_for
    return request.remote_addr or "unknown"


def _cleanup_login_attempts(client_ip: str) -> None:
    """Réinitialise les tentatives lorsque le verrou lié à une IP a expiré."""
    record = ip_login_attempts.get(client_ip)
    if not record:
        return

    lock_until = record.get("lock_until")
    if lock_until and datetime.utcnow() >= lock_until:
        ip_login_attempts.pop(client_ip, None)


def _get_lock_remaining(client_ip: str) -> timedelta | None:
    """Retourne le temps restant avant la fin du verrouillage pour une IP donnée."""
    record = ip_login_attempts.get(client_ip)
    if not record:
        return None

    lock_until = record.get("lock_until")
    if not lock_until:
        return None

    remaining = lock_until - datetime.utcnow()
    if remaining.total_seconds() <= 0:
        ip_login_attempts.pop(client_ip, None)
        return None

    return remaining


def user_can_access_event(user: Utilisateur, evenement: Evenement | None) -> bool:
    """Détermine si l'utilisateur a le droit d'accéder à l'évènement fourni."""

    if evenement is None or user is None:
        return False

    if getattr(user, "is_admin", False):
        return True

    role = (getattr(user, "role", "") or "").lower()
    if role == "codep":
        return True

    if evenement.createur_id and evenement.createur_id == getattr(user, "id", None):
        return True

    try:
        return evenement in user.evenements
    except Exception:
        return False


def _build_lock_context(client_ip: str) -> dict:
    """Construit le contexte de verrouillage pour une IP donnée."""
    remaining = _get_lock_remaining(client_ip)
    if not remaining:
        return {
            "lock_active": False,
            "lock_message": None,
            "lock_minutes": 0,
            "lock_seconds": 0,
            "lock_remaining_seconds": 0,
        }

    total_seconds = int(max(remaining.total_seconds(), 0))
    minutes, seconds = divmod(total_seconds, 60)
    message = (
        "Trop de tentatives infructueuses. Le formulaire est bloqué pendant 5 minutes. "
        f"Réessayez dans {minutes} minute(s) et {seconds} seconde(s)."
    )

    return {
        "lock_active": True,
        "lock_message": message,
        "lock_minutes": minutes,
        "lock_seconds": seconds,
        "lock_remaining_seconds": total_seconds,
    }


@main_bp.route("/", methods=["GET", "POST"])
@limiter.limit("10 per minute", methods=["POST"], error_message="Trop de tentatives de connexion, veuillez patienter avant de réessayer.")
def login():
    client_ip = _get_client_ip()
    _cleanup_login_attempts(client_ip)
    context = _build_lock_context(client_ip)
    context["hide_broadcast"] = True

    if request.method == "GET" and "user_id" in session:
        return redirect(url_for("main_bp.evenement_new"))

    if request.method == "POST":
        if context["lock_active"]:
            context["hide_broadcast"] = True
            return render_template("login.html", **context)

        nom_utilisateur = request.form.get("username", "").strip()
        mot_de_passe = request.form.get("password", "")

        user = Utilisateur.query.filter_by(nom_utilisateur=nom_utilisateur).first()
        password_ok = user.check_password(mot_de_passe) if user else False
        if user and password_ok and user.actif:
            session["user_id"] = user.id
            log_action("login_success", "utilisateur", user.id)
            ip_login_attempts.pop(client_ip, None)
            return redirect(url_for("main_bp.evenement_new"))
        else:
            reason = "inactive" if (user and password_ok and not user.actif) else "invalid_credentials"
            log_action(
                "login_failed",
                "utilisateur",
                user.id if user else None,
                extra=json.dumps(
                    {
                        "username": nom_utilisateur,
                        "ip": client_ip,
                        "reason": reason,
                    },
                    ensure_ascii=False,
                ),
            )

            if reason == "inactive":
                flash("Votre compte est désactivé. Contactez un administrateur.", "danger")
                context["hide_broadcast"] = True
                return render_template("login.html", **context)

            record = ip_login_attempts.setdefault(
                client_ip,
                {"count": 0},
            )
            record["count"] = int(record.get("count", 0)) + 1

            if record["count"] >= LOGIN_LOCK_THRESHOLD:
                record["lock_until"] = datetime.utcnow() + LOGIN_LOCK_DURATION
                context = _build_lock_context(client_ip)
                context["hide_broadcast"] = True
                if context["lock_message"]:
                    flash(context["lock_message"], "danger")
            else:
                remaining_attempts = LOGIN_LOCK_THRESHOLD - record["count"]
                flash(
                    f"Nom d'utilisateur ou mot de passe invalide. "
                    f"Il vous reste {remaining_attempts} tentative(s) avant le blocage.",
                    "danger",
                )
            context = _build_lock_context(client_ip)
            context["hide_broadcast"] = True

    return render_template("login.html", **context)
# 🔓 Déconnexion
@main_bp.route("/logout")
def logout():
    log_action("logout")
    session.clear()
    return redirect(url_for("main_bp.login"))

# 📋 Création + sélection d’un événement
@main_bp.route("/evenement/new", methods=["GET", "POST"])
@login_required
def evenement_new():
    user = get_current_user()

    if request.method == "POST":
        # 🔒 Restriction stricte à admin ou codep
        if not user.is_admin and user.role != "codep":
            flash("⛔ Vous n’avez pas l’autorisation de créer un évènement.", "danger")
            return redirect(url_for("main_bp.evenement_new"))

        nom_evt = request.form["nom_evt"]
        type_evt = request.form["type_evt"]
        adresse = request.form["adresse"]
        statut = request.form["statut"]

        # Génération du numéro d'évènement
        last_evt = Evenement.query.order_by(Evenement.id.desc()).first()
        next_id = last_evt.id + 1 if last_evt else 1
        numero_evt = str(next_id).zfill(8)

        # Création de l'évènement
        nouvel_evt = Evenement(
            numero=numero_evt,
            nom=nom_evt,
            type_evt=type_evt,
            adresse=adresse,
            statut=statut,
            createur_id=user.id,
            date_ouverture=datetime.utcnow()
        )

        db.session.add(nouvel_evt)
        db.session.commit()

        # Association du créateur à l'évènement
        if nouvel_evt not in user.evenements:
            user.evenements.append(nouvel_evt)
            db.session.commit()

        flash("✅ Évènement créé avec succès.", "success")
        return redirect(url_for("main_bp.dashboard", evenement_id=nouvel_evt.id))

    # 🔁 Méthode GET
    if user.is_admin or user.role == "codep":
        evenements = (
            Evenement.query.filter_by(archived=False)
            .order_by(Evenement.date_ouverture.desc())
            .all()
        )
    else:
        evenements = [
            evt
            for evt in sorted(
                user.evenements,
                key=lambda evt: evt.date_ouverture or datetime.min,
                reverse=True,
            )
            if not getattr(evt, "archived", False)
        ]

    statuts_disponibles = sorted({evt.statut for evt in evenements if evt.statut})

    return render_template(
        "evenement_new.html",
        user=user,
        evenements=evenements,
        statuts_disponibles=statuts_disponibles,
        broadcast_emojis=BROADCAST_ALLOWED_EMOJIS,
        broadcast_levels=BROADCAST_ALLOWED_LEVELS,
        broadcast_default_emoji=BROADCAST_DEFAULT_EMOJI,
        broadcast_default_level=BROADCAST_DEFAULT_LEVEL,
        broadcast_level_labels=BROADCAST_LEVEL_LABELS,
    )




@main_bp.route("/notifications/broadcast", methods=["POST"])
@login_required
def create_broadcast():
    user = get_current_user()
    role = (user.role or "").lower() if user else ""
    if not (user and (user.is_admin or role == "codep")):
        abort(403)

    action = (request.form.get("action") or "create").strip().lower()

    if action == "clear":
        updated = BroadcastNotification.query.filter_by(is_active=True).update({"is_active": False})
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
            flash("Impossible de retirer la notification active.", "danger")
            return redirect(url_for("main_bp.evenement_new"))

        if updated:
            log_action("broadcast_cleared")
            flash("La notification active a été désactivée.", "info")
        else:
            flash("Aucune notification active à retirer.", "info")
        return redirect(url_for("main_bp.evenement_new"))

    message = (request.form.get("message") or "").strip()
    if not message:
        flash("Le message de notification ne peut pas être vide.", "danger")
        return redirect(url_for("main_bp.evenement_new"))

    if len(message) > 280:
        message = message[:280].rstrip()

    emoji = (request.form.get("emoji") or BROADCAST_DEFAULT_EMOJI).strip()
    if emoji not in BROADCAST_ALLOWED_EMOJIS:
        emoji = BROADCAST_DEFAULT_EMOJI

    level = (request.form.get("level") or BROADCAST_DEFAULT_LEVEL).strip().lower()
    if level not in BROADCAST_ALLOWED_LEVELS:
        level = BROADCAST_DEFAULT_LEVEL

    BroadcastNotification.query.filter_by(is_active=True).update({"is_active": False})
    notification = BroadcastNotification(
        message=message,
        created_by_id=user.id,
        is_active=True,
        emoji=emoji,
        level=level,
    )
    db.session.add(notification)

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        flash("Impossible d'enregistrer la notification. Merci de réessayer.", "danger")
    else:
        log_action("broadcast_created")
        timeout_seconds = current_app.config.get(
            "BROADCAST_AUTO_CLEAR_SECONDS",
            BROADCAST_AUTO_CLEAR_SECONDS,
        )
        schedule_broadcast_expiration(notification.id, delay=timeout_seconds)
        flash(
            f"Notification envoyée à tous les utilisateurs connectés. Elle sera retirée automatiquement dans {timeout_seconds} secondes.",
            "success",
        )

    return redirect(url_for("main_bp.evenement_new"))


@main_bp.route("/notifications/broadcast/status", methods=["GET"])
@login_required
def get_broadcast_status():
    active = (
        BroadcastNotification.query.filter_by(is_active=True)
        .order_by(BroadcastNotification.created_at.desc())
        .first()
    )
    timeout_seconds = current_app.config.get(
        "BROADCAST_AUTO_CLEAR_SECONDS",
        BROADCAST_AUTO_CLEAR_SECONDS,
    )
    return jsonify(
        {
            "active": active.as_payload() if active else None,
            "autoClearSeconds": timeout_seconds,
        }
    )


@main_bp.route("/evenement/<int:evenement_id>/dashboard")
@login_required
def dashboard(evenement_id):
    session["evenement_id"] = evenement_id
    user = get_current_user()

    evenement = Evenement.query.get(evenement_id)
    if not evenement or not user_can_access_event(user, evenement):
        flash("⛔ Vous n’avez pas accès à cet évènement.", "danger")
        return redirect(url_for("main_bp.evenement_new"))

    fiches = FicheImplique.query.filter_by(evenement_id=evenement.id).all()

    humains = [f for f in fiches if not getattr(f, "est_animal", False)]
    animaux = [f for f in fiches if getattr(f, "est_animal", False)]

    def _is_present(fiche: FicheImplique) -> bool:
        return (fiche.statut or "").strip().lower() == "présent"

    nb_present = sum(1 for f in humains if _is_present(f))
    nb_total = len(humains)
    nb_present_animaux = sum(1 for f in animaux if _is_present(f))
    nb_total_animaux = len(animaux)

    peut_modifier_statut = (
        user.is_admin or
        user.role == "codep" or
        evenement.createur_id == user.id or
        (user.role == "responsable" and user in evenement.utilisateurs)
    )

    return render_template(
        "dashboard.html",
        user=user,
        evenement=evenement,
        fiches=fiches,
        nb_present=nb_present,
        nb_total=nb_total,
        nb_present_animaux=nb_present_animaux,
        nb_total_animaux=nb_total_animaux,
        peut_modifier_statut=peut_modifier_statut,
        competence_colors=COMPETENCE_COLORS
    )




def _format_duration(delta: timedelta | None) -> str:
    if not delta:
        return "—"

    total_seconds = int(delta.total_seconds())
    if total_seconds < 0:
        total_seconds = 0

    days, remainder = divmod(total_seconds, 86400)
    hours, remainder = divmod(remainder, 3600)
    minutes, _ = divmod(remainder, 60)

    parts: list[str] = []
    if days:
        parts.append(f"{days} jour{'s' if days > 1 else ''}")
    if hours:
        parts.append(f"{hours} h")
    if minutes:
        parts.append(f"{minutes} min")

    if not parts:
        return "moins d'une minute"

    return " ".join(parts)


def _build_panorama_data(evenement: Evenement) -> dict[str, typing.Any]:
    fiches = FicheImplique.query.filter_by(evenement_id=evenement.id).all()

    now_utc = datetime.utcnow()

    def _is_present(fiche: FicheImplique) -> bool:
        return (fiche.statut or "").strip().lower() == "présent"

    def _is_sorti(fiche: FicheImplique) -> bool:
        return (fiche.statut or "").strip().lower() == "sorti"

    total_personnes = sum(1 for f in fiches if not getattr(f, "est_animal", False))
    total_animaux = sum(1 for f in fiches if getattr(f, "est_animal", False))

    personnes_presentes = sum(1 for f in fiches if not getattr(f, "est_animal", False) and _is_present(f))
    animaux_presents = sum(1 for f in fiches if getattr(f, "est_animal", False) and _is_present(f))

    personnes_sorties = sum(1 for f in fiches if not getattr(f, "est_animal", False) and _is_sorti(f))
    animaux_sortis = sum(1 for f in fiches if getattr(f, "est_animal", False) and _is_sorti(f))

    durations: list[timedelta] = []
    for fiche in fiches:
        arrivee = fiche.heure_arrivee
        sortie = fiche.heure_sortie or now_utc
        if arrivee:
            try:
                delta = sortie - arrivee
                if isinstance(delta, timedelta):
                    durations.append(delta)
            except Exception:
                continue

    avg_presence = None
    if durations:
        avg_seconds = sum(d.total_seconds() for d in durations) / len(durations)
        avg_presence = timedelta(seconds=avg_seconds)

    competence_counts: Counter[str] = Counter()
    for fiche in fiches:
        raw = (fiche.competences or "").strip()
        if not raw:
            continue
        normalized = unicodedata.normalize("NFKC", raw)
        for item in re.split(r"[,;/\\n|]+", normalized):
            canonical = _canonicalize_competence_label(item)
            if canonical:
                competence_counts[canonical] += 1

    competence_summary = sorted(
        (
            {"label": label, "count": count}
            for label, count in competence_counts.items()
        ),
        key=lambda entry: (-entry["count"], entry["label"].lower()),
    )

    recherches: list[dict[str, typing.Any]] = []
    for fiche in fiches:
        details: list[str] = []
        if fiche.recherche_personne:
            for segment in re.split(r"[\\n;]+", fiche.recherche_personne):
                seg = segment.strip()
                if seg:
                    details.append(seg)
        if fiche.numero_recherche:
            details.append(f"Numéro: {fiche.numero_recherche}")
        if details:
            identite = " ".join(filter(None, [fiche.prenom, fiche.nom])).strip()
            if not identite:
                identite = fiche.numero or f"Fiche #{fiche.id}"
            recherches.append(
                {
                    "identite": identite,
                    "statut": fiche.statut or "",
                    "details": details,
                }
            )

    recherches.sort(key=lambda entry: entry["identite"].lower())

    tickets = Ticket.query.filter_by(evenement_id=evenement.id).all()
    tickets_open = sum(1 for t in tickets if (t.status or "").strip().lower() == "ouvert")
    tickets_in_progress = sum(1 for t in tickets if (t.status or "").strip().lower() == "en cours")

    date_ouverture_locale = getattr(evenement, "date_ouverture_locale", None)
    fonctionnement = None
    if evenement.date_ouverture:
        try:
            fonctionnement = now_utc - evenement.date_ouverture
        except Exception:
            fonctionnement = None

    if date_ouverture_locale:
        try:
            date_ouverture_iso = date_ouverture_locale.isoformat()
        except Exception:
            date_ouverture_iso = None
    else:
        date_ouverture_iso = None

    stats = {
        "personnes_presentes": personnes_presentes,
        "personnes_total": total_personnes,
        "animaux_presents": animaux_presents,
        "animaux_total": total_animaux,
        "personnes_sorties": personnes_sorties,
        "animaux_sortis": animaux_sortis,
        "avg_presence": _format_duration(avg_presence),
        "tickets_open": tickets_open,
        "tickets_in_progress": tickets_in_progress,
        "tickets_total": tickets_open + tickets_in_progress,
        "date_ouverture": date_ouverture_iso,
        "date_ouverture_txt": date_ouverture_locale.strftime("%d/%m/%Y %H:%M") if date_ouverture_locale else "—",
        "fonctionnement": _format_duration(fonctionnement),
        "statut": evenement.statut or "—",
    }

    event_info = {
        "id": evenement.id,
        "nom": evenement.nom or "",
        "adresse": evenement.adresse or "",
        "numero": evenement.numero or "",
        "statut": stats["statut"],
    }

    try:
        import pytz  # type: ignore
        paris = pytz.timezone("Europe/Paris")
    except Exception:
        paris = None

    if date_ouverture_locale and date_ouverture_locale.tzinfo:
        generated_at = datetime.now(date_ouverture_locale.tzinfo)
    elif paris:
        generated_at = datetime.now(paris)
    else:
        generated_at = datetime.utcnow()

    return {
        "event": event_info,
        "stats": stats,
        "competence_summary": competence_summary,
        "recherches": recherches,
        "total_fiches": len(fiches),
        "generated_at": generated_at,
    }


@main_bp.route("/evenement/<int:evenement_id>/panorama")
@login_required
def evenement_panorama(evenement_id: int):
    user = get_current_user()
    evenement = Evenement.query.get_or_404(evenement_id)

    if not user_can_access_event(user, evenement):
        flash("⛔ Vous n’avez pas accès à cet évènement.", "danger")
        return redirect(url_for("main_bp.evenement_new"))

    data = _build_panorama_data(evenement)

    return render_template(
        "evenement_panorama.html",
        user=user,
        evenement=evenement,
        stats=data["stats"],
        competence_summary=data["competence_summary"],
        recherches=data["recherches"],
        total_fiches=data["total_fiches"],
        generated_at=data["generated_at"],
    )


@main_bp.route("/evenement/<int:evenement_id>/panorama_json", methods=["GET"])
@login_required
def evenement_panorama_json(evenement_id: int):
    user = get_current_user()
    evenement = Evenement.query.get_or_404(evenement_id)

    if not user_can_access_event(user, evenement):
        return json_nocache({"error": "unauthorized"}, 403)

    data = _build_panorama_data(evenement)
    payload = {**data}
    generated_at = payload.get("generated_at")
    if isinstance(generated_at, datetime):
        payload["generated_at"] = generated_at.strftime("%d/%m/%Y %H:%M")
    else:
        payload["generated_at"] = None

    return json_nocache(payload)









# 🔁 Sélection d’un événement existant
@main_bp.route("/evenement/select", methods=["POST"])
@login_required
def select_evenement():
    user = get_current_user()
    evt_id = request.form.get("evenement_id")

    if evt_id:
        session["evenement_id"] = int(evt_id)  # 🧠 on stocke dans la session
        return redirect(url_for("main_bp.dashboard", evenement_id=int(evt_id)))
    else:
        flash("Veuillez sélectionner un événement.", "warning")
        return redirect(url_for("main_bp.evenement_new"))





# ➕ Création fiche impliqué (NOUVELLE VERSION)
@main_bp.route("/fiche/new", methods=["GET", "POST"])
@login_required
def fiche_new():
    user = get_current_user()
    evenement_id = session.get("evenement_id")

    if not evenement_id:
        flash("⛔ Aucun évènement actif. Veuillez d'abord accéder à un évènement.", "danger")
        return redirect(url_for("main_bp.evenement_new"))

    evenement = Evenement.query.get(evenement_id)
    if not evenement or not user_can_access_event(user, evenement):
        flash("⛔ Vous n’avez pas accès à cet évènement.", "danger")
        return redirect(url_for("main_bp.evenement_new"))

    fiche_type = (request.values.get("type_fiche") or request.args.get("type") or "humain").strip().lower()
    if fiche_type not in {"humain", "animal"}:
        fiche_type = "humain"

    if request.method == "POST":
        # --- Heure d'arrivée envoyée par le front: "YYYY-MM-DD HH:MM:SS"
        heure_js_str = (request.form.get("heure_arrivee_js") or "").strip()
        try:
            heure_arrivee = datetime.strptime(heure_js_str, "%Y-%m-%d %H:%M:%S")
        except Exception:
            # fallback si vide ou format inattendu
            heure_arrivee = datetime.utcnow()

        # --- Numérotation automatique locale à l’évènement
        last_fiche_evt = (
            FicheImplique.query
            .filter_by(evenement_id=evenement.id)
            .order_by(FicheImplique.id.desc())
            .first()
        )
        next_local = 1
        if last_fiche_evt and last_fiche_evt.numero:
            try:
                last_parts = last_fiche_evt.numero.split("-")
                if len(last_parts) == 2:
                    next_local = int(last_parts[1]) + 1
            except ValueError:
                pass
        numero = f"{str(evenement.id).zfill(3)}-{str(next_local).zfill(4)}"

        if fiche_type == "animal":
            animal_nom = (request.form.get("animal_nom") or "").strip()
            if not animal_nom:
                flash("Le nom de l’animal est obligatoire.", "danger")
                return redirect(request.url)

            animal_espece = (request.form.get("animal_espece") or "").strip()
            if len(animal_espece) > 120:
                flash("L’espèce de l’animal ne peut pas dépasser 120 caractères.", "danger")
                return redirect(request.url)

            animal_particularites = (request.form.get("animal_particularites") or "").strip()
            if len(animal_particularites) > 200:
                flash("Le champ ‘Particularités’ ne peut pas dépasser 200 caractères.", "danger")
                return redirect(request.url)

            animal_notes = (request.form.get("animal_notes") or "").strip()
            if len(animal_notes) > 200:
                flash("Le champ ‘Notes complémentaires’ ne peut pas dépasser 200 caractères.", "danger")
                return redirect(request.url)

            referent_id_raw = (request.form.get("animal_referent_id") or "").strip()
            referent = None
            if referent_id_raw:
                try:
                    referent_id = int(referent_id_raw)
                except ValueError:
                    flash("La personne sélectionnée n’est pas valide.", "danger")
                    return redirect(request.url)

                referent = (
                    FicheImplique.query
                    .filter(
                        FicheImplique.id == referent_id,
                        FicheImplique.evenement_id == evenement.id,
                        FicheImplique.est_animal.is_(False),
                    )
                    .first()
                )
                if not referent:
                    flash("La personne sélectionnée n’existe plus.", "danger")
                    return redirect(request.url)

            fiche = FicheImplique(
                numero=numero,
                nom=animal_nom,
                prenom="",
                adresse=None,
                telephone=None,
                personne_a_prevenir=None,
                tel_personne_a_prevenir=None,
                recherche_personne=None,
                difficultes=animal_particularites,
                competences="",
                est_animal=True,
                humain=False,
                type_fiche="animal",
                animal_espece=animal_espece or None,
                animal_details=animal_notes or None,
                referent_humain=referent,
                statut="présent",
                heure_arrivee=heure_arrivee,
                date_naissance=None,
                utilisateur_id=user.id,
                evenement_id=evenement.id,
                autres_informations=animal_notes or None,
                code_sinus=None,
            )

            db.session.add(fiche)
            db.session.commit()

            try:
                add_timeline(
                    fiche.id,
                    user.id,
                    "Création de la fiche animal",
                    "create",
                )
                db.session.commit()
            except Exception:
                db.session.rollback()

            flash(f"✅ Fiche animal n°{numero} créée pour l’évènement en cours.", "success")
            return redirect(url_for("main_bp.dashboard", evenement_id=evenement.id, focus="animal"))

        # --- Date de naissance (obligatoire) pour les humains
        date_naissance_str = (request.form.get("date_naissance") or "").strip()
        if not date_naissance_str:
            flash("La date de naissance est obligatoire.", "danger")
            return redirect(request.url)
        try:
            date_naissance = datetime.strptime(date_naissance_str, "%Y-%m-%d").date()
        except ValueError:
            flash("Format de date de naissance invalide.", "danger")
            return redirect(request.url)
        if date_naissance > date.today():
            flash("La date de naissance ne peut pas être dans le futur.", "danger")
            return redirect(request.url)

        # --- Champs de base
        nom = (request.form.get("nom") or "").strip()
        prenom = (request.form.get("prenom") or "").strip()
        adresse = (request.form.get("adresse") or "").strip()
        telephone = (request.form.get("telephone") or "").strip()
        personne_a_prevenir = (request.form.get("personne_a_prevenir") or "").strip()
        tel_personne_a_prevenir = (request.form.get("tel_personne_a_prevenir") or "").strip()
        recherche_personne = (request.form.get("recherche_personne") or "").strip()
        difficulte = (request.form.get("difficulte") or "").strip()
        numero_recherche = (request.form.get("numero_recherche") or "").strip()

        code_sinus = (request.form.get("code_sinus") or "").strip()
        if len(code_sinus) > 30:
            flash("Le Code Sinus ne doit pas dépasser 30 caractères.", "danger")
            return redirect(request.url)

        selected_comps = request.form.getlist("competences")
        if "Autre" in selected_comps:
            autre_txt = (request.form.get("competence_autre") or "").strip()
            if not autre_txt:
                flash("Merci de préciser l’autre compétence (20 caractères max).", "danger")
                return redirect(request.url)
            if len(autre_txt) > 20:
                flash("La compétence 'Autre' ne doit pas dépasser 20 caractères.", "danger")
                return redirect(request.url)
            selected_comps = [c for c in selected_comps if c != "Autre"]
            if autre_txt not in selected_comps:
                selected_comps.append(autre_txt)
        if len(selected_comps) > 4:
            flash("⛔ Vous ne pouvez sélectionner que 4 compétences maximum.", "danger")
            return redirect(request.url)
        competences = ",".join(selected_comps)

        autres_infos = (request.form.get("autres_informations") or "").strip()
        if len(autres_infos) > 200:
            flash("Le champ « Autres informations » ne peut pas dépasser 200 caractères.", "danger")
            return redirect(request.url)

        fiche = FicheImplique(
            numero=numero,
            nom=nom,
            prenom=prenom,
            adresse=adresse,
            telephone=telephone,
            personne_a_prevenir=personne_a_prevenir,
            tel_personne_a_prevenir=tel_personne_a_prevenir,
            recherche_personne=recherche_personne,
            difficultes=difficulte,
            competences=competences,
            est_animal=False,
            humain=True,
            type_fiche="humain",
            numero_recherche=numero_recherche,
            statut="présent",
            heure_arrivee=heure_arrivee,
            date_naissance=date_naissance,
            utilisateur_id=user.id,
            evenement_id=evenement.id,
            autres_informations=autres_infos,
            code_sinus=code_sinus if hasattr(FicheImplique, "code_sinus") else None,
        )

        db.session.add(fiche)
        db.session.commit()

        try:
            add_timeline(fiche.id, user.id, "Création de la fiche", "create")
            db.session.commit()
        except Exception:
            db.session.rollback()

        flash(f"✅ Fiche n°{numero} créée pour l’évènement en cours.", "success")
        return redirect(url_for("main_bp.dashboard", evenement_id=evenement.id))

    # --- GET : prévisualisation du prochain numéro
    last_fiche_evt = (
        FicheImplique.query
        .filter_by(evenement_id=evenement.id)
        .order_by(FicheImplique.id.desc())
        .first()
    )
    next_local = 1
    if last_fiche_evt and last_fiche_evt.numero:
        try:
            last_parts = last_fiche_evt.numero.split("-")
            if len(last_parts) == 2:
                next_local = int(last_parts[1]) + 1
        except ValueError:
            pass

    numero_prevu = f"{str(evenement.id).zfill(3)}-{str(next_local).zfill(4)}"

    return render_template(
        "fiche_new.html",
        user=user,
        numero_prevu=numero_prevu,
        competences_list=COMPETENCES_CAI,
        fiche_type=fiche_type,
        referent_lookup_url=url_for("main_bp.fiche_referent_lookup", evenement_id=evenement.id),
    )



    




########################################################

@main_bp.route("/admin/utilisateurs")
@login_required
def admin_utilisateurs():
    user = get_current_user()

    if not user.is_admin and user.role != "codep":
        flash("⛔ Accès refusé : vous n’avez pas les droits pour gérer les utilisateurs.", "danger")
        return redirect(url_for("main_bp.evenement_new"))

    try:
        page = max(int(request.args.get("page", 1)), 1)
    except (TypeError, ValueError):
        page = 1
    per_page = 25
    pagination = (
        Utilisateur.query.order_by(Utilisateur.nom.asc(), Utilisateur.prenom.asc())
        .paginate(page=page, per_page=per_page, error_out=False)
    )

    return render_template(
        "admin_utilisateurs.html",
        utilisateurs=pagination.items,
        pagination=pagination,
        user=user,
        provisional_lifetime_days=PROVISIONAL_ACCOUNT_LIFETIME.days,
    )





################################################################


@main_bp.route("/admin/utilisateurs/<int:utilisateur_id>/toggle", methods=["POST"])
@login_required
def admin_toggle_utilisateur(utilisateur_id: int):
    user = get_current_user()

    if not (user.is_admin or user.role == "codep"):
        abort(403)

    target = Utilisateur.query.get_or_404(utilisateur_id)
    if target.id == user.id:
        flash("Vous ne pouvez pas désactiver votre propre compte.", "warning")
        return redirect(url_for("main_bp.admin_utilisateurs", page=request.args.get("page", 1)))

    target.actif = not bool(target.actif)
    db.session.commit()

    log_action(
        "user_toggle_active",
        "utilisateur",
        target.id,
        extra=json.dumps({"actif": target.actif, "by": user.id}, ensure_ascii=False),
    )

    flash(
        f"Compte {'activé' if target.actif else 'désactivé'} pour {target.nom or target.nom_utilisateur}.",
        "success",
    )
    return redirect(url_for("main_bp.admin_utilisateurs", page=request.args.get("page", 1)))





################################################################


@main_bp.route("/admin/utilisateur/create", methods=["GET", "POST"])
@login_required
def utilisateur_create():
    user = get_current_user()
    if not (user.is_admin or user.role == "codep"):
        flash("Accès refusé", "danger")
        return redirect(url_for("main_bp.dashboard"))

    from app.models import Evenement
    all_evenements = Evenement.query.all()

    if request.method == "POST":
        nom = request.form["nom"].strip()
        nom_utilisateur = request.form["nom_utilisateur"].strip()
        role = (request.form["role"].strip() or "").lower()
        type_utilisateur = request.form["type_utilisateur"].strip()
        password = request.form["password"]
        evenement_ids = request.form.getlist("evenements")
        wants_admin = bool(request.form.get("is_admin"))

        if not user.is_admin:
            if wants_admin:
                flash("Seul un administrateur peut créer un autre administrateur.", "danger")
                return redirect(url_for("main_bp.utilisateur_create"))
            if role == "codep":
                flash("Un compte CODEP ne peut être créé que par un administrateur.", "danger")
                return redirect(url_for("main_bp.utilisateur_create"))

        existing = Utilisateur.query.filter_by(nom_utilisateur=nom_utilisateur).first()
        if existing:
            flash("Nom d'utilisateur déjà utilisé.", "danger")
            return redirect(url_for("main_bp.utilisateur_create"))

        created_now = datetime.utcnow()
        new_user = Utilisateur(
            nom=nom,
            nom_utilisateur=nom_utilisateur,
            role=role,
            type_utilisateur=type_utilisateur,
            is_admin=wants_admin if user.is_admin else False,
            created_at=created_now,
        )
        new_user.set_password(password)

        if new_user.is_provisional:
            new_user.provisional_expires_at = created_now + PROVISIONAL_ACCOUNT_LIFETIME

        for evt_id in evenement_ids:
            evt = Evenement.query.get(int(evt_id))
            if evt:
                new_user.evenements.append(evt)

        db.session.add(new_user)
        db.session.commit()
        flash("Utilisateur créé avec succès", "success")
        return redirect(url_for("main_bp.admin_utilisateurs"))

    return render_template(
        "utilisateur_form.html",
        utilisateur=None,
        all_evenements=all_evenements,
        mode="create",
        current_user=user,
    )


###########################################

    
@main_bp.route("/admin/utilisateur/edit/<int:id>", methods=["GET", "POST"])
@login_required
def utilisateur_edit(id):
    user = get_current_user()
    if not (user.is_admin or user.role == "codep"):
        flash("Accès refusé", "danger")
        return redirect(url_for("main_bp.dashboard"))

    utilisateur = Utilisateur.query.get_or_404(id)
    from app.models import Evenement
    all_evenements = Evenement.query.all()

    if not user.is_admin and utilisateur.id != user.id and (utilisateur.is_admin or utilisateur.role == "codep"):
        flash("Seul un administrateur peut modifier ce compte.", "danger")
        return redirect(url_for("main_bp.admin_utilisateurs"))

    if request.method == "POST":
        role = (request.form["role"].strip() or "").lower()
        type_utilisateur = request.form["type_utilisateur"].strip()
        wants_admin = bool(request.form.get("is_admin"))
        password = request.form["password"]

        if not user.is_admin:
            if role == "codep" and utilisateur.role != "codep":
                flash("Seul un administrateur peut attribuer le rôle CODEP.", "danger")
                return redirect(url_for("main_bp.utilisateur_edit", id=id))
            if wants_admin != utilisateur.is_admin:
                flash("Seul un administrateur peut modifier les droits administrateur.", "danger")
                return redirect(url_for("main_bp.utilisateur_edit", id=id))
        else:
            utilisateur.is_admin = wants_admin

        previous_type = utilisateur.type_utilisateur

        utilisateur.nom = request.form["nom"].strip()
        utilisateur.nom_utilisateur = request.form["nom_utilisateur"].strip()
        utilisateur.role = role or utilisateur.role
        utilisateur.type_utilisateur = type_utilisateur or utilisateur.type_utilisateur

        if utilisateur.is_provisional:
            if utilisateur.provisional_expires_at is None or (previous_type or "").lower() != "provisoire":
                utilisateur.provisional_expires_at = datetime.utcnow() + PROVISIONAL_ACCOUNT_LIFETIME
        else:
            utilisateur.provisional_expires_at = None

        if password:
            if user.is_admin:
                utilisateur.set_password(password)
            else:
                if utilisateur.id == user.id:
                    utilisateur.set_password(password)
                elif not utilisateur.is_admin and utilisateur.role != "codep":
                    utilisateur.set_password(password)
                else:
                    flash("Seul un administrateur peut modifier le mot de passe de ce compte.", "danger")
                    return redirect(url_for("main_bp.utilisateur_edit", id=id))

        utilisateur.evenements = []
        for evt_id in request.form.getlist("evenements"):
            evt = Evenement.query.get(int(evt_id))
            if evt:
                utilisateur.evenements.append(evt)

        db.session.commit()
        flash("Utilisateur mis à jour.", "success")
        return redirect(url_for("main_bp.admin_utilisateurs"))

    return render_template(
        "utilisateur_form.html",
        utilisateur=utilisateur,
        all_evenements=all_evenements,
        mode="edit",
        current_user=user,
    )




@main_bp.route("/admin/utilisateur/delete/<int:id>")
@login_required
def utilisateur_delete(id):
    user = get_current_user()
    if not (user.is_admin or user.role in ["responsable", "codep"]):
        flash("Accès refusé.", "danger")
        return redirect(url_for("main_bp.dashboard"))

    utilisateur = Utilisateur.query.get_or_404(id)
    if not user.is_admin and user.role == "codep" and (utilisateur.is_admin or utilisateur.role == "codep"):
        flash("Seul un administrateur peut supprimer ce compte.", "danger")
        return redirect(url_for("main_bp.admin_utilisateurs"))

    db.session.delete(utilisateur)
    db.session.commit()
    flash("Utilisateur supprimé.", "info")
    return redirect(url_for("main_bp.admin_utilisateurs"))


# 🔍 Détail d’une fiche impliqué
@main_bp.route("/fiche/<int:id>")
@login_required
def fiche_detail(id):
    user = get_current_user()
    fiche = FicheImplique.query.get_or_404(id)

    if not user_can_access_event(user, fiche.evenement):
        flash("⛔ Vous n'avez pas accès à cette fiche.", "danger")
        return redirect(url_for("main_bp.evenement_new"))

    entries = fiche.timeline_entries.order_by(TimelineEntry.created_at.desc()).all()
    return render_template("fiche_detail.html", fiche=fiche, user=user, entries=entries)

# ✏️ Modification d’une fiche impliqué
from datetime import datetime

# ✏️ Modification d’une fiche impliqué (MÀJ)
@main_bp.route("/fiche/edit/<int:id>", methods=["GET", "POST"])
@login_required
def fiche_edit(id):
    user = get_current_user()
    fiche = FicheImplique.query.get_or_404(id)
    role_lower = (user.role or "").lower()

    # Vérification d'accès à l'évènement
    if not user_can_access_event(user, fiche.evenement):
        flash("⛔ Vous n’avez pas accès à cet évènement.", "danger")
        return redirect(url_for("main_bp.evenement_new"))

    if request.method == "POST":
        if fiche.est_animal:
            before = {
                "nom": fiche.nom or "",
                "animal_espece": fiche.animal_espece or "",
                "animal_details": fiche.animal_details or "",
                "difficultes": fiche.difficultes or "",
                "referent_id": fiche.referent_humain_id or "",
                "statut": fiche.statut or "",
            }

            fiche_nom = (request.form.get("animal_nom") or "").strip()
            if not fiche_nom:
                flash("Le nom de l’animal est obligatoire.", "danger")
                return redirect(request.url)

            fiche_espece = (request.form.get("animal_espece") or "").strip()
            if len(fiche_espece) > 120:
                flash("L’espèce de l’animal ne peut pas dépasser 120 caractères.", "danger")
                return redirect(request.url)

            fiche_particularites = (request.form.get("animal_particularites") or "").strip()
            if len(fiche_particularites) > 200:
                flash("Le champ ‘Particularités’ ne peut pas dépasser 200 caractères.", "danger")
                return redirect(request.url)

            fiche_notes = (request.form.get("animal_notes") or "").strip()
            if len(fiche_notes) > 200:
                flash("Le champ ‘Notes complémentaires’ ne peut pas dépasser 200 caractères.", "danger")
                return redirect(request.url)

            referent_raw = (request.form.get("animal_referent_id") or "").strip()
            referent = None
            if referent_raw:
                try:
                    referent_id = int(referent_raw)
                except ValueError:
                    flash("La personne sélectionnée n’est pas valide.", "danger")
                    return redirect(request.url)

                referent = (
                    FicheImplique.query
                    .filter(
                        FicheImplique.id == referent_id,
                        FicheImplique.evenement_id == fiche.evenement_id,
                        FicheImplique.est_animal.is_(False),
                    )
                    .first()
                )
                if not referent:
                    flash("La personne sélectionnée n’existe plus.", "danger")
                    return redirect(request.url)

            statut = (request.form.get("statut") or fiche.statut or "présent").strip()
            fiche.nom = fiche_nom
            fiche.prenom = ""
            fiche.animal_espece = fiche_espece or None
            fiche.animal_details = fiche_notes or None
            fiche.difficultes = fiche_particularites or None
            fiche.autres_informations = fiche_notes or None
            fiche.referent_humain = referent
            fiche.statut = statut or "présent"
            fiche.est_animal = True
            fiche.humain = False
            fiche.type_fiche = "animal"

            db.session.commit()

            after = {
                "nom": fiche.nom or "",
                "animal_espece": fiche.animal_espece or "",
                "animal_details": fiche.animal_details or "",
                "difficultes": fiche.difficultes or "",
                "referent_id": fiche.referent_humain_id or "",
                "statut": fiche.statut or "",
            }

            labels = {
                "nom": "nom",
                "animal_espece": "espèce",
                "animal_details": "notes",
                "difficultes": "particularités",
                "referent_id": "référent humain",
                "statut": "statut",
            }

            changes = []
            for key, lib in labels.items():
                if before.get(key) != after.get(key):
                    old = before.get(key, "") or "—"
                    new = after.get(key, "") or "—"
                    if key == "referent_id":
                        old = str(old) if old else "aucun"
                        new = str(new) if new else "aucun"
                    changes.append(f"{lib}: «{old}» → «{new}»")

            if changes:
                try:
                    add_timeline(
                        fiche.id,
                        user.id,
                        "Modification fiche animal — " + "; ".join(changes[:6]),
                        "update",
                    )
                    db.session.commit()
                except Exception:
                    db.session.rollback()

            flash("✅ Fiche animal mise à jour avec succès.", "success")
            return redirect(url_for("main_bp.dashboard", evenement_id=fiche.evenement.id))

        # ====== Snapshot avant modifications (pour diff) ======
        def _val(v):  # normalise None -> ""
            return "" if v is None else str(v)
        before = {
            "nom": _val(fiche.nom),
            "prenom": _val(fiche.prenom),
            "statut": _val(fiche.statut),
            "difficulte": _val(fiche.difficultes),
            "telephone": _val(fiche.telephone),
            "adresse": _val(fiche.adresse),
            "recherche_personne": _val(fiche.recherche_personne),
            "destination": _val(fiche.destination),
            "moyen_transport": _val(fiche.moyen_transport),
            "personne_a_prevenir": _val(fiche.personne_a_prevenir),
            "numero_pec": _val(getattr(fiche, "numero_pec", "")),
            "tel_personne_a_prevenir": _val(fiche.tel_personne_a_prevenir),
            "code_sinus": _val(getattr(fiche, "code_sinus", "")),
            "competences": _val(fiche.competences),
            "autres_informations": _val(fiche.autres_informations),
            "date_naissance": fiche.date_naissance.strftime("%Y-%m-%d") if fiche.date_naissance else "",
        }

        # ====== Application des modifications ======
        fiche.nom = request.form.get("nom")
        fiche.prenom = request.form.get("prenom")
        fiche.statut = request.form.get("statut")
        fiche.difficultes = request.form.get("difficulte")
        fiche.telephone = request.form.get("telephone")
        fiche.adresse = request.form.get("adresse")
        fiche.recherche_personne = request.form.get("recherche_personne")
        fiche.destination = request.form.get("destination")
        fiche.moyen_transport = request.form.get("moyen_transport")
        fiche.personne_a_prevenir = request.form.get("personne_a_prevenir")
        if user.is_admin or role_lower not in {"technicien", "logisticien"}:
            fiche.numero_pec = request.form.get("numero_pec")
        fiche.tel_personne_a_prevenir = request.form.get("tel_personne_a_prevenir")

        # 🆕 Code Sinus (30 max)
        code_sinus = (request.form.get("code_sinus") or "").strip()
        if len(code_sinus) > 30:
            flash("Le Code Sinus ne doit pas dépasser 30 caractères.", "danger")
            return redirect(request.url)
        if hasattr(fiche, "code_sinus"):
            fiche.code_sinus = code_sinus

        # ✅ Compétences + 'Autre' (20 max) + limite 4
        selected_comps = request.form.getlist("competences")
        if "Autre" in selected_comps:
            autre_txt = (request.form.get("competence_autre") or "").strip()
            if not autre_txt:
                flash("Merci de préciser l’autre compétence (20 caractères max).", "danger")
                return redirect(request.url)
            if len(autre_txt) > 20:
                flash("La compétence 'Autre' ne doit pas dépasser 20 caractères.", "danger")
                return redirect(request.url)
            selected_comps = [c for c in selected_comps if c != "Autre"]
            if autre_txt not in selected_comps:
                selected_comps.append(autre_txt)
        if len(selected_comps) > 4:
            flash("⛔ Vous ne pouvez sélectionner que 4 compétences maximum.", "danger")
            return redirect(request.url)
        fiche.competences = ",".join(selected_comps)

        # ✅ Autres informations (trim + limite 200)
        autres_infos = (request.form.get("autres_informations") or "").strip()
        if len(autres_infos) > 200:
            flash("Le champ « Autres informations » ne peut pas dépasser 200 caractères.", "danger")
            return redirect(request.url)
        fiche.autres_informations = autres_infos

        fiche.est_animal = False
        fiche.humain = True
        fiche.type_fiche = "humain"
        fiche.animal_espece = None
        fiche.animal_details = None
        fiche.referent_humain = None

        # ✅ Conversion de la date au bon format (obligatoire)
        date_str = (request.form.get("date_naissance") or "").strip()
        if not date_str:
            flash("La date de naissance est obligatoire.", "danger")
            return redirect(request.url)
        try:
            parsed_birth = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            flash("⚠️ Format de date invalide.", "danger")
            return redirect(request.url)
        if parsed_birth > date.today():
            flash("La date de naissance ne peut pas être dans le futur.", "danger")
            return redirect(request.url)
        fiche.date_naissance = parsed_birth

        # ====== 1er commit : on valide la mise à jour ======
        db.session.commit()

        # ====== Diff après modifications ======
        after = {
            "nom": _val(fiche.nom),
            "prenom": _val(fiche.prenom),
            "statut": _val(fiche.statut),
            "difficulte": _val(fiche.difficultes),
            "telephone": _val(fiche.telephone),
            "adresse": _val(fiche.adresse),
            "recherche_personne": _val(fiche.recherche_personne),
            "destination": _val(fiche.destination),
            "moyen_transport": _val(fiche.moyen_transport),
            "personne_a_prevenir": _val(fiche.personne_a_prevenir),
            "numero_pec": _val(getattr(fiche, "numero_pec", "")),
            "tel_personne_a_prevenir": _val(fiche.tel_personne_a_prevenir),
            "code_sinus": _val(getattr(fiche, "code_sinus", "")),
            "competences": _val(fiche.competences),
            "autres_informations": _val(fiche.autres_informations),
            "date_naissance": fiche.date_naissance.strftime("%Y-%m-%d") if fiche.date_naissance else "",
        }

        # Champs lisibles pour la timeline (clé → libellé)
        labels = {
            "nom": "nom",
            "prenom": "prénom",
            "statut": "statut",
            "difficulte": "difficultés",
            "telephone": "téléphone",
            "adresse": "adresse",
            "recherche_personne": "recherche personne",
            "destination": "destination",
            "moyen_transport": "moyen transport",
            "personne_a_prevenir": "personne à prévenir",
            "numero_pec": "n° PEC",
            "tel_personne_a_prevenir": "tél. à prévenir",
            "code_sinus": "code sinus",
            "competences": "compétences",
            "autres_informations": "autres informations",
            "date_naissance": "date de naissance",
        }

        changes = []
        for k, lib in labels.items():
            if before.get(k, "") != after.get(k, ""):
                old = before.get(k, "")
                new = after.get(k, "")
                # On compacte un peu les très longues valeurs
                def short(s):
                    s = s or ""
                    return (s[:60] + "…") if len(s) > 60 else s
                changes.append(f"{lib}: «{short(old)}» → «{short(new)}»")

        # Ajout timeline si au moins 1 champ a changé
        if changes:
            try:
                content = "Modification fiche — " + "; ".join(changes[:6])  # on limite à 6 changements pour rester lisible
                if len(changes) > 6:
                    content += f" (+{len(changes)-6} autres)"
                add_timeline(fiche.id, user.id, content, "update")
                db.session.commit()
            except Exception:
                db.session.rollback()

        flash("✅ Fiche mise à jour avec succès.", "success")
        return redirect(url_for("main_bp.dashboard", evenement_id=fiche.evenement.id))

    return render_template(
        "fiche_edit.html",
        fiche=fiche,
        user=user,
        competences_list=COMPETENCES_CAI,
        fiche_type="animal" if fiche.est_animal else "humain",
        referent_lookup_url=url_for("main_bp.fiche_referent_lookup", evenement_id=fiche.evenement_id),
    )






########################################################################

@main_bp.route("/fiche/delete/<int:id>", methods=["POST"])
@login_required
def fiche_delete(id):
    user = get_current_user()
    fiche = FicheImplique.query.get_or_404(id)

    # Doit avoir les droits de rôle
    roles_autorises = {"responsable", "codep"}
    if not (user.is_admin or (user.role or "").lower() in roles_autorises):
        flash("⛔ Suppression réservée à un administrateur, un codep ou un responsable.", "danger")
        return redirect(url_for("main_bp.dashboard", evenement_id=fiche.evenement.id))

    # Doit avoir accès à l'évènement
    if not user_can_access_event(user, fiche.evenement):
        flash("⛔ Vous n’avez pas accès à cet évènement.", "danger")
        return redirect(url_for("main_bp.evenement_new"))

    db.session.delete(fiche)
    db.session.commit()
    flash("🗑️ Fiche supprimée avec succès.", "info")
    return redirect(url_for("main_bp.dashboard", evenement_id=fiche.evenement.id))




################################################################################



@main_bp.route("/fiche/<int:id>/sortie", methods=["POST"])
@login_required
def fiche_sortie(id):
    fiche = FicheImplique.query.get_or_404(id)
    user = get_current_user()

    if not user_can_access_event(user, fiche.evenement):
        flash("⛔ Vous n’avez pas accès à cette fiche.", "danger")
        return redirect(url_for("main_bp.dashboard", evenement_id=fiche.evenement_id))

    # Champs envoyés par la popup
    destination = (request.form.get("destination") or "").strip()
    moyen_transport = (request.form.get("moyen_transport") or "").strip()

    # Mise à jour de la fiche
    if destination:
        fiche.destination = destination
    if moyen_transport:
        fiche.moyen_transport = moyen_transport

    fiche.statut = "sorti"
    fiche.heure_sortie = datetime.utcnow()

    # 1er commit : on valide la mise à jour de la fiche
    db.session.commit()

    # ✅ Timeline : trace de la sortie
    try:
        details = []
        if destination:
            details.append(f"destination={destination}")
        if moyen_transport:
            details.append(f"transport={moyen_transport}")
        content = "Sortie" + (f" ({', '.join(details)})" if details else "")
        add_timeline(fiche.id, user.id, content, "exit")
        db.session.commit()
    except Exception:
        db.session.rollback()

    flash(f"🚪 {fiche.nom} {fiche.prenom} est marqué comme 'sorti'.", "info")
    return redirect(url_for("main_bp.dashboard", evenement_id=fiche.evenement_id))





###############################################################



@main_bp.route("/evenement/<int:evenement_id>/update_statut", methods=["POST"])
@login_required
def update_evenement_statut(evenement_id):
    user = get_current_user()
    evenement = Evenement.query.get_or_404(evenement_id)

    if not user_can_access_event(user, evenement):
        flash("⛔ Accès refusé.", "danger")
        return redirect(url_for("main_bp.evenement_new"))

    new_statut = request.form.get("statut_evt")
    if new_statut:
        evenement.statut = new_statut
        db.session.commit()
        flash("✅ Statut de l’évènement mis à jour.", "success")

    return redirect(url_for("main_bp.dashboard", evenement_id=evenement.id))

#############################################

@main_bp.route("/evenement/<int:evenement_id>/fiches_json", methods=["GET"])
@login_required
def fiches_json(evenement_id):
    user = get_current_user()
    evt = Evenement.query.get_or_404(evenement_id)

    # 🔐 Accès lecture : rattaché à l’évènement OU admin/codep
    if not user_can_access_event(user, evt):
        return json_nocache({"error": "unauthorized"}, 403)

    # Tri stable par id (asc) pour éviter les sauts d’ordre au refresh
    fiches = (
        FicheImplique.query
        .filter_by(evenement_id=evenement_id)
        .order_by(FicheImplique.id.asc())
        .all()
    )

    def fmt(dt):
        try:
            return dt.strftime("%d/%m/%Y %H:%M") if dt else "-"
        except Exception:
            return "-"

    fiches_data = []
    for f in fiches:
        h_arr = getattr(f, "heure_arrivee_locale", None)
        h_out = getattr(f, "heure_sortie_locale", None)
        fiches_data.append({
            "id": f.id,
            "numero": f.numero or "",
            "nom": f.nom or "",
            "prenom": f.prenom or "",
            "statut": f.statut or "",
            "heure_arrivee": fmt(h_arr),
            "heure_sortie": fmt(h_out),
            "destination": f.destination or "",
            "difficultes": f.difficultes or "",
            "competences": f.competences or "",
            "est_animal": bool(getattr(f, "est_animal", False)),
            "type_fiche": (f.type_fiche or ("animal" if getattr(f, "est_animal", False) else "humain")),
            "animal_espece": f.animal_espece or "",
            "referent": (
                {
                    "id": f.referent_humain.id,
                    "label": " ".join(
                        part for part in [(f.referent_humain.prenom or ""), (f.referent_humain.nom or "")] if part
                    ).strip() or (f.referent_humain.numero or ""),
                }
                if getattr(f, "referent_humain", None)
                else None
            ),
        })

    # Métadonnées évènement pour mise à jour du bandeau
    date_ouv_loc = getattr(evt, "date_ouverture_locale", None)
    evt_payload = {
        "id": evt.id,
        "nom": evt.nom or "",
        "adresse": evt.adresse or "",
        "statut": evt.statut or "",
        "date_ouverture": fmt(date_ouv_loc),
    }

    # Compteurs live
    humains = [f for f in fiches if not getattr(f, "est_animal", False)]
    animaux = [f for f in fiches if getattr(f, "est_animal", False)]

    def _is_present(fiche: FicheImplique) -> bool:
        return (fiche.statut or "").strip().lower() == "présent"

    nb_present = sum(1 for f in humains if _is_present(f))
    nb_total = len(humains)
    nb_present_animaux = sum(1 for f in animaux if _is_present(f))
    nb_total_animaux = len(animaux)

    return json_nocache({
        "fiches": fiches_data,
        "nb_present": nb_present,
        "nb_total": nb_total,
        "nb_present_animaux": nb_present_animaux,
        "nb_total_animaux": nb_total_animaux,
        "evenement": evt_payload,
    })


@main_bp.route("/evenement/<int:evenement_id>/referents", methods=["GET"])
@login_required
def fiche_referent_lookup(evenement_id: int):
    user = get_current_user()
    evt = Evenement.query.get_or_404(evenement_id)

    if not user_can_access_event(user, evt):
        return json_nocache({"error": "unauthorized"}, 403)

    term = (request.args.get("term") or "").strip().lower()

    query = (
        FicheImplique.query
        .filter(
            FicheImplique.evenement_id == evenement_id,
            FicheImplique.est_animal.is_(False),
        )
    )

    if term:
        like_term = f"%{term}%"
        query = query.filter(
            or_(
                func.lower(FicheImplique.nom).like(like_term),
                func.lower(FicheImplique.prenom).like(like_term),
                func.lower(FicheImplique.numero).like(like_term),
            )
        )

    fiches = (
        query
        .order_by(func.lower(FicheImplique.prenom).asc(), func.lower(FicheImplique.nom).asc())
        .limit(15)
        .all()
    )

    results = []
    for fiche in fiches:
        label_parts = [fiche.prenom or "", fiche.nom or ""]
        label = " ".join(part for part in label_parts if part).strip() or fiche.numero
        results.append({
            "id": fiche.id,
            "label": label,
            "numero": fiche.numero,
        })

    return json_nocache({"results": results})



#####################################################################

COMPETENCE_COLORS = {
    "Médecin": "#e74c3c",
    "Infirmier": "#3498db",
    "Sapeur-pompier": "#e67e22",
    "SST": "#1abc9c",
    "Psychologue": "#9b59b6",
    "Bénévole": "#34495e",
    "Artisan": "#f39c12",
    "Interprète": "#2ecc71",
    "Logisticien": "#16a085",
    "Conducteur": "#d35400",
    "Agent sécurité": "#2c3e50",
    "Autre": "#7f8c8d"
}



#############################################"

def _styled_table(data):
    table = Table(data, colWidths=[60*mm, 100*mm])
    table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 11),
        ('BACKGROUND', (0,0), (-1,-1), colors.whitesmoke),
        ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.whitesmoke, colors.lightgrey]),
        ('TEXTCOLOR', (0,0), (-1,-1), colors.black),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('INNERGRID', (0,0), (-1,-1), 0.3, colors.grey),
        ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
    ]))
    return table


@main_bp.route("/fiche/<int:id>/pdf")
@login_required
def export_pdf_fiche(id):
    fiche = FicheImplique.query.get_or_404(id)

    tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    tmp_file.close()
    doc = SimpleDocTemplate(tmp_file.name, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=60, bottomMargin=40)

    story = []

    # === STYLES ===
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Titre', fontSize=22, alignment=1, textColor=colors.HexColor("#002f6c"), spaceAfter=20))
    styles.add(ParagraphStyle(name='SectionTitle', fontSize=14, textColor=colors.HexColor("#f58220"), spaceBefore=15, spaceAfter=8, underlineWidth=1))
    styles.add(ParagraphStyle(name='NormalBold', parent=styles['Normal'], fontName='Helvetica-Bold'))

    # === LOGO + TITRE ===
    logo_path = os.path.join("static", "img", "logo-protection-civile.jpg")
    if os.path.exists(logo_path):
        img = Image(logo_path, width=70, height=70)
        img.hAlign = 'CENTER'
        story.append(img)

    story.append(Paragraph("Fiche Impliqué", styles['Titre']))

    # === INFOS PERSO ===
    story.append(Paragraph("Informations personnelles", styles['SectionTitle']))
    data_perso = [
        ["Numéro", fiche.numero],
        ["Nom", fiche.nom],
        ["Prénom", fiche.prenom],
        ["Date de naissance", fiche.date_naissance.strftime('%d/%m/%Y') if fiche.date_naissance else "Non renseignée"],
        ["Nationalité", fiche.nationalite or "Non renseignée"],
        ["Adresse", fiche.adresse or "Non renseignée"],
        ["Téléphone", fiche.telephone or "Non renseigné"],
    ]
    story.append(_styled_table(data_perso))

    # === CONTACT & COORDONNÉES ===
    story.append(Paragraph("Coordonnées et contact", styles['SectionTitle']))
    data_contact = [
        ["Téléphone", fiche.telephone or "Non renseigné"],
        ["Personne à prévenir", fiche.personne_a_prevenir or "Non renseignée"],
        ["Téléphone personne à prévenir", fiche.tel_personne_a_prevenir or "Non renseigné"],
        ["Numéro PEC", fiche.numero_pec or "Non renseigné"],
    ]
    story.append(_styled_table(data_contact))

    # === INFOS HORAIRES & STATUT ===
    story.append(Paragraph("Suivi opérationnel", styles['SectionTitle']))
    data_horaires = [
        ["Statut", fiche.statut or "Non renseigné"],
        ["Heure d’arrivée", fiche.heure_arrivee_locale.strftime('%d/%m/%Y %H:%M') if fiche.heure_arrivee_locale else "Non renseignée"],
        ["Heure de sortie", fiche.heure_sortie_locale.strftime('%d/%m/%Y %H:%M') if fiche.heure_sortie_locale else "Non renseignée"],
        ["Destination", fiche.destination or "Non renseignée"],
        ["Moyen de transport", fiche.moyen_transport or "Non renseigné"],
    ]
    story.append(_styled_table(data_horaires))

    # === INFORMATIONS MÉDICALES / LOGISTIQUES ===
    story.append(Paragraph("Informations complémentaires", styles['SectionTitle']))
    data_supp = [
        ["Code Sinus", fiche.code_sinus or "Non renseigné"],
        ["Recherche une personne", fiche.recherche_personne or "Non"],
        ["Numéro de recherche", fiche.numero_recherche or "Non renseigné"],
        ["Difficultés", fiche.difficultes or "Non renseignées"],
        ["Compétences", fiche.competences or "Non renseignées"],
        ["Effets personnels", fiche.effets_perso or "Non renseignés"],
        ["Est un animal", "Oui" if fiche.est_animal else "Non"],
        ["Est humain", "Oui" if fiche.humain else "Non"],
    ]
    story.append(_styled_table(data_supp))

    # === ÉVÈNEMENT ASSOCIÉ ===
    story.append(Paragraph("Évènement associé", styles['SectionTitle']))
    evenement_data = [
        ["Nom", fiche.evenement.nom if fiche.evenement else "Non renseigné"],
        ["Numéro", fiche.evenement.numero if fiche.evenement else "Non renseigné"],
        ["Adresse", fiche.evenement.adresse if fiche.evenement else "Non renseignée"],
        ["Statut", fiche.evenement.statut if fiche.evenement else "Non renseigné"],
        ["Créateur de la fiche", (fiche.createur.nom if fiche.createur else "Non renseigné")],
    ]
    story.append(_styled_table(evenement_data))

    # === AUTRES INFORMATIONS ===
    if fiche.autres_informations:
        story.append(Paragraph("Autres informations", styles['SectionTitle']))
        story.append(_styled_table([["Notes", fiche.autres_informations]]))

    # === BAGAGES ===
    story.append(Paragraph("Bagages", styles['SectionTitle']))
    try:
        bag_list = sorted(
            [b.numero for b in (fiche.bagages or []) if b and b.numero],
            key=lambda x: x
        )
    except Exception:
        bag_list = []

    bagages_str = ", ".join(bag_list) if bag_list else "Aucun"
    story.append(_styled_table([["Bagages rattachés", bagages_str]]))

    # === HISTORIQUE TIMELINE ===
    timeline_entries = (
        fiche.timeline_entries.order_by(TimelineEntry.created_at.desc()).limit(10).all()
    )
    if timeline_entries:
        import pytz

        story.append(Paragraph("Historique récent", styles['SectionTitle']))
        paris = pytz.timezone("Europe/Paris")
        timeline_rows = []
        for entry in reversed(timeline_entries):
            try:
                ts = entry.created_at.astimezone(paris) if entry.created_at else None
            except Exception:
                ts = entry.created_at
            auteur = entry.user.nom if entry.user and entry.user.nom else (entry.user.nom_utilisateur if entry.user else "")
            timeline_rows.append([
                ts.strftime('%d/%m/%Y %H:%M') if ts else "—",
                f"{auteur or '—'} — {entry.content}",
            ])
        story.append(_styled_table([["Date", "Commentaire"]] + timeline_rows))

    doc.build(story)

    def generate() -> typing.Iterator[bytes]:
        try:
            with open(tmp_file.name, "rb") as f:
                while True:
                    chunk = f.read(8192)
                    if not chunk:
                        break
                    yield chunk
        finally:
            try:
                os.remove(tmp_file.name)
            except OSError:
                pass

    filename = f"fiche_{fiche.numero or fiche.id}.pdf"
    response = Response(stream_with_context(generate()), mimetype="application/pdf")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response



################################################

@main_bp.route("/admin/evenements")
@login_required
def admin_evenements():
    user = get_current_user()

    if not user.is_admin and user.role != "codep":
        flash("⛔ Accès interdit.", "danger")
        return redirect(url_for("main_bp.evenement_new"))

    evenements = (
        Evenement.query.filter_by(archived=False)
        .order_by(Evenement.id.desc())
        .all()
    )
    archived_evenements = (
        Evenement.query.filter_by(archived=True)
        .order_by(Evenement.id.desc())
        .all()
    )
    return render_template(
        "admin_evenements.html",
        evenements=evenements,
        archived_evenements=archived_evenements,
        user=user,
    )


####################################

@main_bp.route('/evenement/<int:evenement_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_evenement(evenement_id):
    user = get_current_user()
    evenement = Evenement.query.get_or_404(evenement_id)

    if not user.is_admin and user.role != "codep" and evenement.createur_id != user.id:
        flash("⛔ Accès interdit.", "danger")
        return redirect(url_for("main_bp.admin_evenements"))

    if request.method == "POST":
        evenement.nom = request.form["nom"]
        evenement.adresse = request.form["adresse"]
        evenement.type_evt = request.form["type"]
        evenement.statut = request.form["statut"]
        db.session.commit()
        flash("✅ Évènement mis à jour.", "success")
        return redirect(url_for("main_bp.admin_evenements"))

    return render_template("edit_evenement.html", evenement=evenement, user=user)

#########################################


@main_bp.route("/evenements/<int:evenement_id>/supprimer", methods=["POST"])
@login_required
def delete_evenement(evenement_id):
    user = get_current_user()  # ✅ au lieu de current_user
    evt = Evenement.query.get_or_404(evenement_id)

    # 🔐 Vérifie si l'utilisateur est admin OU le créateur (codep)
    if not (user.is_admin or user.role == "codep" or evt.createur_id == user.id):
        abort(403)

    # 🧹 Supprime les fiches impliquées et leurs dépendances (bagages, timeline…)
    from .models import Ticket

    for fiche in list(evt.impliques):
        for bagage in list(getattr(fiche, "bagages", []) or []):
            db.session.delete(bagage)
        db.session.delete(fiche)

    # 🧹 Supprime aussi les tickets associés
    for ticket in list(evt.tickets):
        db.session.delete(ticket)

    # 🧹 Supprime les actualités liées à l'évènement
    for news in list(evt.news):
        db.session.delete(news)

    # 🧹 Supprime les liens de partage liés à l'évènement
    for share_link in list(evt.share_links):
        db.session.delete(share_link)

    # 🗑 Supprime l'évènement
    db.session.delete(evt)
    db.session.commit()

    flash("✅ L’évènement et ses fiches ont été supprimés.", "success")
    return redirect(url_for("main_bp.evenement_new"))


@main_bp.route("/evenements/<int:evenement_id>/archiver", methods=["POST"])
@login_required
def archive_evenement(evenement_id: int):
    user = get_current_user()
    evt = Evenement.query.get_or_404(evenement_id)

    if not (user.is_admin or user.role == "codep"):
        abort(403)

    if not evt.archived:
        evt.archived = True
        evt.archived_at = datetime.utcnow()
        db.session.commit()
        log_action(
            "evenement_archived",
            "evenement",
            evt.id,
            extra=json.dumps({"by": user.id}, ensure_ascii=False),
        )
        flash("📦 L’évènement a été archivé.", "info")
    else:
        flash("L’évènement est déjà archivé.", "warning")

    return redirect(url_for("main_bp.admin_evenements"))


@main_bp.route("/evenements/<int:evenement_id>/restaurer", methods=["POST"])
@login_required
def restore_evenement(evenement_id: int):
    user = get_current_user()
    evt = Evenement.query.get_or_404(evenement_id)

    if not (user.is_admin or user.role == "codep"):
        abort(403)

    if evt.archived:
        evt.archived = False
        evt.archived_at = None
        db.session.commit()
        log_action(
            "evenement_restored",
            "evenement",
            evt.id,
            extra=json.dumps({"by": user.id}, ensure_ascii=False),
        )
        flash("✅ L’évènement a été restauré.", "success")
    else:
        flash("L’évènement est déjà actif.", "info")

    return redirect(url_for("main_bp.admin_evenements"))



###################################################



@main_bp.route("/evenement/<int:evenement_id>/export/pdf")
@login_required
def export_evenement_fiches_pdf(evenement_id):
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    import io
    import pytz

    evenement = Evenement.query.get_or_404(evenement_id)
    fiches = FicheImplique.query.filter_by(evenement_id=evenement_id).all()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    elements = []

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='CenterTitle', alignment=1, fontSize=18, spaceAfter=20))
    styles.add(ParagraphStyle(name='SubHeader', textColor=colors.orange, fontSize=14, spaceAfter=10))

    elements.append(Paragraph("Fiches Impliqués – Évènement", styles['CenterTitle']))
    elements.append(Paragraph("Informations sur l’évènement", styles['SubHeader']))

    # Date locale
    def convertir_heure_locale(dt_utc):
        if not dt_utc:
            return "Non renseignée"
        paris = pytz.timezone("Europe/Paris")
        return dt_utc.astimezone(paris).strftime("%d/%m/%Y %H:%M")

    infos_evt = [
        ["Nom", evenement.nom],
        ["Numéro", evenement.numero],
        ["Adresse", evenement.adresse],
        ["Statut", evenement.statut],
        ["Type", evenement.type_evt],
        ["Date d'ouverture", convertir_heure_locale(evenement.date_ouverture)]
    ]
    table_evt = Table(infos_evt, hAlign='LEFT', colWidths=[4*cm, 12*cm])
    table_evt.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
    ]))
    elements.append(table_evt)
    elements.append(Spacer(1, 20))

    elements.append(Paragraph("Liste des fiches impliquées", styles['SubHeader']))

    header = [
        "Nom", "Prénom", "Naissance", "Nationalité", "Statut",
        "Téléphone", "Adresse", "Compétences", "Destination", "Effets perso"
    ]
    data = [header]

    for f in fiches:
        row = [
            f.nom or "-",
            f.prenom or "-",
            f.date_naissance.strftime("%d/%m/%Y") if f.date_naissance else "-",
            f.nationalite or "-",
            f.statut or "-",
            f.telephone or "-",
            f.adresse or "-",
            f.competences or "-",
            f.destination or "-",
            f.effets_perso or "-",
        ]
        data.append(row)

    table_fiches = Table(data, repeatRows=1, hAlign='LEFT')
    table_fiches.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey)
    ]))
    elements.append(table_fiches)

    doc.build(elements)
    buffer.seek(0)

    return send_file(buffer, as_attachment=True, download_name=f"evenement_{evenement.numero}_fiches.pdf", mimetype='application/pdf')



###########################################



@main_bp.route("/fiche/<int:fiche_id>/bagages/ajouter", methods=["POST"])
@login_required
def fiche_bagages_ajouter(fiche_id):
    user = get_current_user()
    fiche = FicheImplique.query.get_or_404(fiche_id)

    if not user_can_access_event(user, fiche.evenement):
        flash("⛔ Vous n’avez pas accès à cet évènement.", "danger")
        return redirect(url_for("main_bp.evenement_new"))

    role = (user.role or "").lower()
    if not (user.is_admin or role in {"technicien", "responsable", "codep"}):
        flash("⛔ Vous n’êtes pas autorisé à modifier les bagages.", "danger")
        return redirect(url_for("main_bp.dashboard", evenement_id=fiche.evenement_id))

    raw = (request.form.get("numeros") or "").strip()
    # Découpe par virgules, espaces, points-virgules et retours à la ligne
    nouveaux = [t.strip() for t in re.split(r"[\s,;]+", raw) if t.strip()]
    # Dédoublonner en conservant l’ordre
    uniques, vus = [], set()
    for t in nouveaux:
        if t not in vus:
            uniques.append(t)
            vus.add(t)
    nouveaux_set = set(uniques)

    # État actuel
    existants = Bagage.query.filter_by(fiche_id=fiche.id).all()
    existants_map = {b.numero: b for b in existants}
    existants_set = set(existants_map.keys())

    # Diff
    a_supprimer = existants_set - nouveaux_set
    a_ajouter = nouveaux_set - existants_set

    deja_autre_fiche = []
    ajoutes = []
    supprimes = []

    # Unicité au niveau évènement : un numéro ne peut pas être utilisé par une autre fiche du même centre
    if a_ajouter:
        doublons_centre = {
            b.numero: b
            for b in Bagage.query.filter(
                Bagage.evenement_id == fiche.evenement_id,
                Bagage.numero.in_(list(a_ajouter))
            ).all()
        }
    else:
        doublons_centre = {}

    # Ajouter
    for num in a_ajouter:
        autre = doublons_centre.get(num)
        if autre and autre.fiche_id != fiche.id:
            deja_autre_fiche.append(num)
            continue
        db.session.add(Bagage(numero=num, fiche_id=fiche.id, evenement_id=fiche.evenement_id))
        ajoutes.append(num)

    # Supprimer
    for num in a_supprimer:
        db.session.delete(existants_map[num])
        supprimes.append(num)

    db.session.commit()

    # Feedback
    parts = []
    if ajoutes: parts.append(f"Ajouté: {', '.join(sorted(ajoutes))}")
    if supprimes: parts.append(f"Supprimé: {', '.join(sorted(supprimes))}")
    if deja_autre_fiche: parts.append(f"En conflit (déjà utilisés par une autre fiche): {', '.join(sorted(deja_autre_fiche))}")
    flash(" | ".join(parts) if parts else "Aucune modification.", "success" if (ajoutes or supprimes) else "info")

    return redirect(url_for("main_bp.dashboard", evenement_id=fiche.evenement_id))


##################################

@main_bp.route("/fiche/<int:fiche_id>/bagages_json", methods=["GET"])
@login_required
def fiche_bagages_json(fiche_id):
    user = get_current_user()
    fiche = FicheImplique.query.get_or_404(fiche_id)

    # 🔐 Accès lecture : rattaché à l’évènement OU admin/codep
    if not user_can_access_event(user, fiche.evenement):
        return json_nocache({"error": "unauthorized"}, 403)

    numeros = [
        b.numero for b in Bagage.query
        .filter_by(fiche_id=fiche.id)
        .order_by(Bagage.id.asc())
        .all()
        if b.numero
    ]
    return json_nocache({"fiche_id": fiche.id, "numero_fiche": fiche.numero, "numeros": numeros})


###########################################################

@main_bp.route("/evenement/<int:evenement_id>/export/csv")
@login_required
def export_evenement_fiches_csv(evenement_id):
    from datetime import datetime
    import pytz
    from flask import redirect, url_for, flash
    from openpyxl import Workbook
    try:
        from openpyxl.cell import WriteOnlyCell  # openpyxl ≥ 3.1
    except ImportError:  # pragma: no cover - compat older versions
        from openpyxl.writer.write_only import WriteOnlyCell
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    user = get_current_user()
    evenement = Evenement.query.get_or_404(evenement_id)

    # Permissions : admin, codep, responsable rattaché
    if not (
        user.is_admin
        or user.role == "codep"
        or (user.role == "responsable" and user in evenement.utilisateurs)
    ):
        flash("⛔ Accès refusé pour l’export.", "danger")
        return redirect(url_for("main_bp.dashboard", evenement_id=evenement.id))

    fiches_query = (
        FicheImplique.query
        .filter_by(evenement_id=evenement.id)
        .order_by(FicheImplique.id.asc())
        .yield_per(200)
    )

    nb_total = db.session.query(func.count(FicheImplique.id)).filter_by(evenement_id=evenement.id).scalar() or 0
    nb_present = (
        db.session.query(func.count(FicheImplique.id))
        .filter(FicheImplique.evenement_id == evenement.id)
        .filter(func.lower(FicheImplique.statut) == "présent")
        .scalar()
        or 0
    )
    nb_sorti = (
        db.session.query(func.count(FicheImplique.id))
        .filter(FicheImplique.evenement_id == evenement.id)
        .filter(func.lower(FicheImplique.statut) == "sorti")
        .scalar()
        or 0
    )

    bagages_map: dict[int, list[str]] = {}
    for bagage in Bagage.query.filter_by(evenement_id=evenement.id).yield_per(200):
        bagages_map.setdefault(bagage.fiche_id, []).append(bagage.numero)

    # Timezone Paris
    paris = pytz.timezone("Europe/Paris")
    def to_paris_dt(dt):
        if not dt: return None
        try:
            return dt.astimezone(paris).replace(tzinfo=None)
        except Exception:
            return None

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Fiches Impliqués")
    if len(wb.worksheets) > 1:
        wb.remove(wb.worksheets[0])

    BLEU = "002F6C"
    ORANGE = "F58220"
    GRIS_LIGNE = "E9EDF3"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=BLEU)
    title_font = Font(bold=True, color="FFFFFF", size=14)
    title_fill = PatternFill("solid", fgColor=ORANGE)
    border_thin = Border(
        left=Side(style="thin", color=GRIS_LIGNE),
        right=Side(style="thin", color=GRIS_LIGNE),
        top=Side(style="thin", color=GRIS_LIGNE),
        bottom=Side(style="thin", color=GRIS_LIGNE),
    )
    wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center")

    def make_cell(value, font=None, fill=None, alignment=None, border=None, number_format=None):
        cell = WriteOnlyCell(ws, value=value)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
        if number_format:
            cell.number_format = number_format
        return cell

    ws.append([
        make_cell("📋 Export Fiches Impliqués — Protection Civile", font=title_font, fill=title_fill, alignment=center, border=border_thin)
    ])
    ws.append([])

    summary_rows = [
        ("Évènement", evenement.nom or ""),
        ("Numéro", evenement.numero or ""),
        ("Adresse", evenement.adresse or ""),
        ("Statut", evenement.statut or ""),
        ("Type", evenement.type_evt or ""),
        ("Ouverture", to_paris_dt(evenement.date_ouverture)),
        ("Présents", nb_present),
        ("Total / Sortis", f"{nb_total} / {nb_sorti}"),
    ]
    for label, value in summary_rows:
        cells = [
            make_cell(label, font=Font(bold=True, color=BLEU), alignment=wrap_left, border=border_thin),
        ]
        if isinstance(value, datetime):
            cells.append(make_cell(value, alignment=wrap_left, border=border_thin, number_format="DD/MM/YYYY HH:MM"))
        else:
            cells.append(make_cell(value, alignment=wrap_left, border=border_thin))
        ws.append(cells)

    ws.append([])

    headers = [
        "Numéro",
        "Code Sinus",
        "Nom",
        "Prénom",
        "Date de naissance",
        "Téléphone",
        "Adresse",
        "Statut",
        "Heure d’arrivée",
        "Heure de sortie",
        "Destination",
        "Moyen de transport",
        "Recherche personne",
        "N° recherche",
        "Personne à prévenir",
        "Tél. à prévenir",
        "Difficultés",
        "Compétences",
        "Bagages",
        "Autres informations",
    ]
    ws.append([make_cell(h, font=header_font, fill=header_fill, alignment=center, border=border_thin) for h in headers])

    for fiche in fiches_query:
        bagages_txt = ", ".join(sorted(bagages_map.get(fiche.id, []))) if bagages_map.get(fiche.id) else ""
        naissance = fiche.date_naissance if isinstance(fiche.date_naissance, datetime) else fiche.date_naissance
        arrivee = fiche.heure_arrivee_locale.replace(tzinfo=None) if getattr(fiche, "heure_arrivee_locale", None) else to_paris_dt(getattr(fiche, "heure_arrivee", None))
        sortie = fiche.heure_sortie_locale.replace(tzinfo=None) if getattr(fiche, "heure_sortie_locale", None) else to_paris_dt(getattr(fiche, "heure_sortie", None))

        row_values = [
            fiche.numero or "",
            getattr(fiche, "code_sinus", "") or "",
            fiche.nom or "",
            fiche.prenom or "",
            naissance,
            fiche.telephone or "",
            fiche.adresse or "",
            fiche.statut or "",
            arrivee,
            sortie,
            fiche.destination or "",
            fiche.moyen_transport or "",
            fiche.recherche_personne or "",
            getattr(fiche, "numero_recherche", "") or "",
            fiche.personne_a_prevenir or "",
            fiche.tel_personne_a_prevenir or "",
            fiche.difficultes or "",
            fiche.competences or "",
            bagages_txt,
            fiche.autres_informations or "",
        ]

        row_cells = []
        for idx, value in enumerate(row_values):
            number_format = None
            if idx == 4 and isinstance(value, datetime):
                number_format = "DD/MM/YYYY"
            if idx in (8, 9) and isinstance(value, datetime):
                number_format = "DD/MM/YYYY HH:MM"
            row_cells.append(make_cell(value, alignment=wrap_left, border=border_thin, number_format=number_format))
        ws.append(row_cells)

    tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_file.close()
    wb.save(tmp_file.name)

    def generate() -> typing.Iterator[bytes]:
        try:
            with open(tmp_file.name, "rb") as fh:
                while True:
                    chunk = fh.read(8192)
                    if not chunk:
                        break
                    yield chunk
        finally:
            try:
                os.remove(tmp_file.name)
            except OSError:
                pass

    filename = f"evenement_{evenement.numero or evenement.id}_fiches.xlsx"
    response = Response(
        stream_with_context(generate()),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response



##################################################################


def can_manage_sharing(user):
    return user.is_admin or user.role in {"codep", "responsable"}

# ===== Création d’un lien de partage (affiche le token UNE fois) =====
@main_bp.route("/evenement/<int:evenement_id>/share/create", methods=["POST"])
@limiter.limit("5 per minute", error_message="Trop de créations de lien de partage, réessayez ultérieurement.")
@login_required
def create_share_link(evenement_id):
    user = get_current_user()
    evt = Evenement.query.get_or_404(evenement_id)

    if not can_manage_sharing(user):
        abort(403)

    import secrets, hashlib
    token = secrets.token_urlsafe(24)  # clair
    token_hash = hashlib.sha256(token.encode()).hexdigest()
    label = (request.form.get("label") or "").strip()
    if label:
        label = label[:120]

    link = ShareLink(
        token=token,                # ✅ sauvegarde du clair
        token_hash=token_hash,      # ✅ sauvegarde du hash
        evenement_id=evt.id,
        created_by=user.id,
        label=label or None,
    )
    db.session.add(link)
    db.session.commit()

    flash("🔗 Lien de partage créé.", "success")
    return redirect(url_for("main_bp.autorite_dashboard_manage", evenement_id=evt.id))



# ===== Gestion des liens : page opérateur =====
@main_bp.route("/evenement/<int:evenement_id>/autorite", methods=["GET"])
@login_required
def autorite_dashboard_manage(evenement_id):
    import pytz
    paris = pytz.timezone("Europe/Paris")

    user = get_current_user()
    evt = Evenement.query.get_or_404(evenement_id)

    if not (user.is_admin or user.role in {"codep", "responsable"} or evt.createur_id == user.id):
        flash("⛔ Accès refusé.", "danger")
        return redirect(url_for("main_bp.dashboard", evenement_id=evenement_id))

    links = ShareLink.query.filter_by(evenement_id=evenement_id).order_by(ShareLink.created_at.desc()).all()

    share_metrics: dict[int, dict] = {}
    if links:
        link_ids = [link.id for link in links]
        logs = (
            ShareLinkAccessLog.query.filter(ShareLinkAccessLog.share_link_id.in_(link_ids))
            .order_by(ShareLinkAccessLog.accessed_at.desc())
            .all()
        )
        logs_by_link: dict[int, list[ShareLinkAccessLog]] = {}
        for log in logs:
            logs_by_link.setdefault(log.share_link_id, []).append(log)

        for link in links:
            entries = logs_by_link.get(link.id, [])
            try:
                last_access = entries[0].accessed_at.astimezone(paris) if entries else None
            except Exception:
                last_access = entries[0].accessed_at if entries else None
            recent_history = []
            for entry in entries[:10]:
                try:
                    accessed_at = entry.accessed_at.astimezone(paris)
                except Exception:
                    accessed_at = entry.accessed_at
                recent_history.append(
                    {
                        "at": accessed_at,
                        "ip": entry.ip,
                        "user_agent": entry.user_agent,
                    }
                )
            share_metrics[link.id] = {
                "total": len(entries),
                "unique_ips": len({entry.ip for entry in entries if entry.ip}),
                "last_access": last_access,
                "recent": recent_history,
            }

    # Récupération des actus avec conversion locale
    all_news = EventNews.query.filter_by(evenement_id=evenement_id).order_by(
        EventNews.priority.asc(), EventNews.created_at.desc()
    ).all()
    for n in all_news:
        if n.created_at:
            try:
                n.created_at_local = n.created_at.astimezone(paris)
            except Exception:
                n.created_at_local = n.created_at

    one_time_token = request.args.get("token")
    return render_template(
        "autorite_dashboard.html",
        user=user,
        evenement=evt,
        links=links,
        share_metrics=share_metrics,
        manage=True,
        one_time_token=one_time_token,
        all_news=all_news
    )



# ===== Révocation par ID (pas par token qu’on ne stocke pas) =====
@main_bp.route("/share/<int:link_id>/revoke", methods=["POST"])
@login_required
def revoke_share_link(link_id):
    user = get_current_user()
    link = ShareLink.query.get_or_404(link_id)

    if not (user.is_admin or user.role in {"codep", "responsable"} or link.evenement.createur_id == user.id):
        abort(403)

    link.revoked = True
    db.session.commit()
    flash("❌ Lien révoqué.", "info")
    return redirect(url_for("main_bp.autorite_dashboard_manage", evenement_id=link.evenement_id))


# 🔓 Vue PUBLIQUE (read-only) par token, pour coller à ton template:
@main_bp.route("/autorite/<token>", methods=["GET"])
@limiter.limit("60 per minute", error_message="Trop de requêtes sur ce lien public, merci de patienter.")
def autorite_share_public(token):
    # On cherche le lien (même s'il est révoqué, on veut distinguer les cas)
    link = ShareLink.query.filter_by(token=token).first()

    # Lien introuvable ou révoqué -> page dédiée
    if not link or link.revoked:
        resp = render_template("autorite_share_invalid.html", hide_broadcast=True)
        # 410 Gone = ressource n'est plus disponible (meilleur qu'un 404 ici)
        return resp, 410, {
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
        }

    # Lien valide -> on charge l'évènement
    ev = Evenement.query.get_or_404(link.evenement_id)

    try:
        access_log = ShareLinkAccessLog(
            share_link_id=link.id,
            ip=_get_client_ip(),
            user_agent=(request.headers.get("User-Agent") or "")[:255],
            referer=(request.referrer or "")[:255],
        )
        db.session.add(access_log)
        db.session.commit()
    except Exception:
        db.session.rollback()

    # On réutilise le même template; manage=False masque la gestion
    # public_token=token pour que le JS appelle /autorite_json?token=...
    return render_template(
        "autorite_dashboard.html",
        evenement=ev,
        manage=False,
        links=None,
        public_token=token,
        hide_broadcast=True,
    )

#####################################################

@main_bp.route("/evenement/<int:evenement_id>/autorite_json", methods=["GET"])
def autorite_json(evenement_id):
    import pytz
    paris = pytz.timezone("Europe/Paris")

    token = request.args.get("token")

    if token:
        link = ShareLink.query.filter_by(token=token, revoked=False, evenement_id=evenement_id).first()
        if not link:
            return jsonify({"error": "invalid_or_revoked_token"}), 410, {
                "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
                "Pragma": "no-cache",
                "Expires": "0",
            }
        ev = Evenement.query.get_or_404(evenement_id)
    else:
        if not current_user.is_authenticated:
            abort(401)
        ev = Evenement.query.get_or_404(evenement_id)

    fiches = (
        db.session.query(FicheImplique)
        .filter_by(evenement_id=evenement_id)
        .all()
    )

    def _is_status(fiche: FicheImplique, statut: str) -> bool:
        return (fiche.statut or "").strip().lower() == statut

    humains = [f for f in fiches if not getattr(f, "est_animal", False)]
    animaux = [f for f in fiches if getattr(f, "est_animal", False)]

    nb_total = len(humains)
    nb_present = sum(1 for f in humains if _is_status(f, "présent"))
    nb_sorti = sum(1 for f in humains if _is_status(f, "sorti"))
    nb_animaux_total = len(animaux)
    nb_animaux_present = sum(1 for f in animaux if _is_status(f, "présent"))

    # 🔥 ACTUS actives avec heure locale Paris
    news_q = (EventNews.query
              .filter_by(evenement_id=evenement_id, is_active=True)
              .order_by(EventNews.priority.asc(), EventNews.created_at.desc()))
    news_items = []
    for n in news_q.limit(12).all():
        created_local = None
        if n.created_at:
            try:
                created_local = n.created_at.astimezone(paris).strftime("%d/%m/%Y %H:%M")
            except Exception:
                created_local = n.created_at.strftime("%d/%m/%Y %H:%M")
        news_items.append({
            "id": n.id,
            "message": n.message,
            "priority": n.priority,
            "icon": n.icon,
            "created_at": created_local
        })

    date_ouverture = getattr(ev, "date_ouverture", None)
    if date_ouverture:
        try:
            date_str = date_ouverture.astimezone(paris).strftime("%d/%m/%Y %H:%M")
        except Exception:
            try:
                date_str = date_ouverture.replace(tzinfo=timezone.utc).astimezone(paris).strftime("%d/%m/%Y %H:%M")
            except Exception:
                date_str = date_ouverture.strftime("%d/%m/%Y %H:%M")
    else:
        date_str = ""

    temps_fonctionnement = "—"
    if date_ouverture:
        try:
            start = date_ouverture if date_ouverture.tzinfo else date_ouverture.replace(tzinfo=timezone.utc)
            now_utc = datetime.now(timezone.utc)
            delta = now_utc - start.astimezone(timezone.utc)
            total_minutes = max(int(delta.total_seconds() // 60), 0)
            jours, minutes_restantes = divmod(total_minutes, 1440)
            heures, minutes = divmod(minutes_restantes, 60)
            parts = []
            if jours:
                parts.append(f"{jours}j")
            parts.append(f"{heures}h")
            parts.append(f"{minutes:02d}m")
            temps_fonctionnement = " ".join(parts)
        except Exception:
            temps_fonctionnement = "—"

    return jsonify({
        "evenement": {
            "id": ev.id,
            "nom": ev.nom,
            "adresse": ev.adresse,
            "statut": ev.statut,
            "date_ouverture": date_str,
        },
        "stats": {
            "nb_total": nb_total,
            "nb_present": nb_present,
            "nb_sorti": nb_sorti,
            "nb_animaux_present": nb_animaux_present,
            "nb_animaux_total": nb_animaux_total,
            "temps_fonctionnement": temps_fonctionnement,
        },
        "news": news_items
    }), 200, {
        "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
        "Pragma": "no-cache",
        "Expires": "0",
    }








#######################



def has_ticket_rights(user):
    return bool(
        user.is_admin or
        (user.role or "").lower() in {"codep", "responsable", "logisticien"}
    )


# ===== TICKETS =====

@main_bp.route("/evenement/<int:evenement_id>/tickets")
@login_required
def tickets_board(evenement_id):
    user = get_current_user()
    evenement = Evenement.query.get_or_404(evenement_id)
    users = evenement.utilisateurs  # ou ta logique de sélection
    
    can_manage = user.is_admin or user.role in ["codep", "responsable", "logisticien"]
    
    users_data = [
        {"id": u.id, "nom": u.nom, "role": u.role}
        for u in users
    ]
    
    return render_template(
        "tickets_board.html",
        evenement=evenement,
        users=users,
        users_data=users_data,  # ✅ on passe ici
        can_manage=can_manage,
        user=user
    )

################################################

# --- Création d'une actu depuis la page de gestion ---
@main_bp.route("/evenement/<int:evenement_id>/news", methods=["POST"])
@login_required
def create_event_news(evenement_id):
    ev = Evenement.query.get_or_404(evenement_id)

    msg = (request.form.get("message") or "").strip()
    priority = int(request.form.get("priority") or 3)
    icon = (request.form.get("icon") or "fa-circle-info").strip()  # classe fa-... sans le 'fa-solid' devant

    if not msg:
        flash("Le message est obligatoire.", "danger")
        return redirect(url_for("main_bp.autorite_dashboard_manage", evenement_id=evenement_id))

    news = EventNews(
        evenement_id=ev.id,
        created_by=(current_user.id if current_user.is_authenticated else None),
        message=msg,
        priority=max(1, min(priority, 3)),
        icon=icon,
        is_active=True,
    )
    db.session.add(news)
    db.session.commit()
    flash("Actualité ajoutée.", "success")
    return redirect(url_for("main_bp.autorite_dashboard_manage", evenement_id=evenement_id))

# --- Basculer actif/inactif ---
@main_bp.route("/news/<int:news_id>/toggle", methods=["POST"])
@login_required
def toggle_event_news(news_id):
    news = EventNews.query.get_or_404(news_id)
    news.is_active = not news.is_active
    db.session.commit()
    flash("Actualité mise à jour.", "success")
    return redirect(url_for("main_bp.autorite_dashboard_manage", evenement_id=news.evenement_id))

# --- Supprimer ---
@main_bp.route("/news/<int:news_id>/delete", methods=["POST"])
@login_required
def delete_event_news(news_id):
    news = EventNews.query.get_or_404(news_id)
    eid = news.evenement_id
    db.session.delete(news)
    db.session.commit()
    flash("Actualité supprimée.", "success")
    return redirect(url_for("main_bp.autorite_dashboard_manage", evenement_id=eid))

############################

@main_bp.route("/evenement/<int:evenement_id>/tickets_json", methods=["GET"])
@login_required
def tickets_json(evenement_id):
    user = get_current_user()
    evt = Evenement.query.get_or_404(evenement_id)

    # Accès lecture : toute personne rattachée à l’évènement ou admin
    if not user_can_access_event(user, evt):
        return json_nocache({"error": "forbidden"}, 403)

    tickets = (
        Ticket.query
        .filter_by(evenement_id=evt.id)
        .order_by(Ticket.created_at.desc())
        .all()
    )
    return json_nocache({"tickets": [t.to_dict() for t in tickets]})
############################


@main_bp.route("/tickets/create", methods=["POST"])
@login_required
def ticket_create():
    user = get_current_user()
    if not has_ticket_rights(user):
        flash("⛔ Vous n'êtes pas autorisé à créer des tickets.", "danger")
        return redirect(url_for("main_bp.evenement_new"))

    evenement_id = int(request.form.get("evenement_id"))
    evt = Evenement.query.get_or_404(evenement_id)
    if not user_can_access_event(user, evt):
        flash("⛔ Accès refusé.", "danger")
        return redirect(url_for("main_bp.evenement_new"))

    title = (request.form.get("title") or "").strip()
    if not title:
        flash("Le titre est obligatoire.", "danger")
        return redirect(url_for("main_bp.tickets_board", evenement_id=evenement_id))

    t = Ticket(
        evenement_id=evenement_id,
        created_by_id=user.id,
        title=title,
        description=(request.form.get("description") or "").strip(),
        status=(request.form.get("status") or "Ouvert"),
        priority=(request.form.get("priority") or "Normal"),
        category=(request.form.get("category") or "Logistique"),
        assigned_to_id=(int(request.form.get("assigned_to_id")) if request.form.get("assigned_to_id") else None),
    )
    db.session.add(t)
    db.session.commit()
    flash("🎫 Ticket créé.", "success")
    return redirect(url_for("main_bp.tickets_board", evenement_id=evenement_id))


@main_bp.route("/tickets/<int:ticket_id>/update", methods=["POST"])
@login_required
def ticket_update(ticket_id):
    user = get_current_user()
    t = Ticket.query.get_or_404(ticket_id)
    evt = Evenement.query.get_or_404(t.evenement_id)
    if not user_can_access_event(user, evt):
        return jsonify({"error": "unauthorized"}), 403
    if not has_ticket_rights(user):
        return jsonify({"error": "forbidden"}), 403

    # champs autorisés à modifier
    t.status = request.form.get("status", t.status)
    t.priority = request.form.get("priority", t.priority)
    t.category = request.form.get("category", t.category)
    t.description = request.form.get("description", t.description)
    if "assigned_to_id" in request.form:
        val = request.form.get("assigned_to_id")
        t.assigned_to_id = int(val) if val else None

    db.session.commit()
    return jsonify({"ok": True})


@main_bp.route("/tickets/<int:ticket_id>/delete", methods=["POST"])
@login_required
def ticket_delete(ticket_id):
    user = get_current_user()
    t = Ticket.query.get_or_404(ticket_id)
    evt = Evenement.query.get_or_404(t.evenement_id)

    if not user_can_access_event(user, evt) or not has_ticket_rights(user):
        flash("⛔ Suppression non autorisée.", "danger")
        return redirect(url_for("main_bp.evenement_new"))

    db.session.delete(t)
    db.session.commit()
    flash("🗑️ Ticket supprimé.", "info")
    return redirect(url_for("main_bp.tickets_board", evenement_id=evt.id))


# === Sauvegarde (téléchargement direct) ===
@main_bp.route("/admin/backup", methods=["GET"])
@limiter.limit("3 per minute", error_message="Trop de demandes de sauvegarde, merci de patienter.")
@login_required
def admin_backup():
    user = get_current_user()
    if not user.is_admin:
        abort(403)
    evenement_ids: list[int] = []
    for raw_id in request.args.getlist("evenement_id"):
        try:
            evenement_ids.append(int(raw_id))
        except (TypeError, ValueError):
            continue

    buf = backup_to_bytesio(evenement_ids or None)
    if evenement_ids:
        suffix = "_".join(str(eid) for eid in sorted(set(evenement_ids)))
        filename = f"backup_evenements_{suffix}.json"
    else:
        filename = "backup.json"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/json"
    )

# === Page de gestion (boutons + formulaire restauration) ===
@main_bp.route("/admin/backup-restore", methods=["GET"])
@login_required
def admin_backup_restore_page():
    user = get_current_user()
    if not user.is_admin:
        abort(403)
    evenements = Evenement.query.order_by(Evenement.nom.asc()).all()
    return render_template("admin_backup_restore.html", user=user, evenements=evenements)

# === Restauration ===
@main_bp.route("/admin/restore", methods=["POST"])
@limiter.limit("1 per minute", error_message="Trop de demandes de restauration, réessayez plus tard.")
@login_required
def admin_restore():

    user = get_current_user()
    if not (user and user.is_admin):
        abort(403)

    # --- Récupération du fichier + option "forcer" ---
    file = request.files.get("backup_file")
    force = request.form.get("force") == "on"

    if not file or file.filename.strip() == "":
        flash("Aucun fichier fourni.", "danger")
        return redirect(url_for("main_bp.admin_backup_restore_page"))

    # --- Lecture JSON ---
    try:
        raw = file.read()
        payload = json.loads(raw.decode("utf-8"))
    except Exception as e:
        flash(f"Fichier invalide (JSON) : {e}", "danger")
        return redirect(url_for("main_bp.admin_backup_restore_page"))

    # --- Réglages SQLite pour limiter les verrous pendant l'opération ---
    try:
        with db.engine.connect() as con:
            con.exec_driver_sql("PRAGMA busy_timeout=10000")
            con.exec_driver_sql("PRAGMA journal_mode=WAL")
    except Exception:
        # Non bloquant si l’on n’est pas sur SQLite
        pass

    # --- Sécurité : refuser si la base n'est pas vide (sauf option 'forcer') ---
    try:
        if not force and not is_db_empty():
            flash("Restauration refusée : la base n’est pas vide (sécurité). Coche « Forcer » pour écraser.", "warning")
            return redirect(url_for("main_bp.admin_backup_restore_page"))
    except Exception as e:
        flash(f"Vérification de base vide impossible : {e}", "danger")
        return redirect(url_for("main_bp.admin_backup_restore_page"))

    # --- Si 'forcer', on efface proprement avant de restaurer ---
    if force:
        try:
            # Ferme toute session en cours (évite 'database is locked')
            db.session.remove()
            wipe_db(preserve_users=True)
        except Exception as e:
            flash(f"Impossible d’effacer la base : {e}", "danger")
            return redirect(url_for("main_bp.admin_backup_restore_page"))

    # --- Injection en base ---
    try:
        bulk_restore(payload)
    except Exception as e:
        # On tente un rollback propre pour laisser l'app dans un état stable
        try:
            db.session.rollback()
        except Exception:
            pass
        flash(f"Erreur pendant la restauration : {e}", "danger")
        return redirect(url_for("main_bp.admin_backup_restore_page"))

    flash("✅ Restauration terminée avec succès.", "success")
    return redirect(url_for("main_bp.admin_backup_restore_page"))


@main_bp.route('/healthz')
@limiter.exempt
def _healthz():
    health = {"status": "ok", "checks": {}}

    try:
        db.session.execute(text("SELECT 1"))
        health["checks"]["database"] = "ok"
    except Exception as exc:
        health["status"] = "error"
        health["checks"]["database"] = str(exc)

    try:
        redis_url = os.getenv("REDIS_URL", "redis://redis:6379/0")
        client = redis.from_url(redis_url)
        client.ping()
        health["checks"]["redis"] = "ok"
    except Exception as exc:
        health["status"] = "error"
        health["checks"]["redis"] = str(exc)

    status_code = 200 if health["status"] == "ok" else 503
    return jsonify(health), status_code


# =====================
# Page d'audit (admin uniquement)
# =====================
@main_bp.route("/admin/logs")
@login_required
def admin_logs():
    user = get_current_user()
    if not getattr(user, "is_admin", False):
        flash("⛔ Accès réservé à l'administrateur.", "danger")
        return redirect(url_for("main_bp.dashboard"))
    page = max(int(request.args.get("page", 1)), 1)
    per_page = max(min(int(request.args.get("per_page", 50)), 200), 10)

    filters = {
        "user_id": request.args.get("user_id", type=int),
        "action": (request.args.get("action") or "").strip(),
        "entity": (request.args.get("entity") or "").strip(),
        "q": (request.args.get("q") or "").strip(),
    }

    query = AuditLog.query
    if filters["user_id"]:
        query = query.filter(AuditLog.user_id == filters["user_id"])
    if filters["action"]:
        query = query.filter(AuditLog.action == filters["action"])
    if filters["entity"]:
        query = query.filter(AuditLog.entity_type == filters["entity"])
    if filters["q"]:
        like = f"%{filters['q']}%"
        query = query.filter(
            or_(
                AuditLog.action.ilike(like),
                AuditLog.entity_type.ilike(like),
                func.cast(AuditLog.entity_id, db.String).ilike(like),
                func.cast(AuditLog.user_id, db.String).ilike(like),
                AuditLog.ip.ilike(like),
                AuditLog.extra.ilike(like),
            )
        )

    query = query.order_by(AuditLog.created_at.desc())
    logs = query.paginate(page=page, per_page=per_page, error_out=False)

    users = Utilisateur.query.order_by(Utilisateur.nom_utilisateur.asc()).all()
    actions = [row[0] for row in db.session.query(AuditLog.action).distinct().order_by(AuditLog.action.asc()) if row[0]]
    entities = [row[0] for row in db.session.query(AuditLog.entity_type).filter(AuditLog.entity_type.isnot(None)).distinct().order_by(AuditLog.entity_type.asc())]

    active_filters = {k: v for k, v in filters.items() if v}

    return render_template(
        "admin_logs.html",
        logs=logs,
        user=user,
        users=users,
        actions=actions,
        entities=entities,
        filters=filters,
        active_filters=active_filters,
        per_page=per_page,
    )


# =====================
# Audit : suppression ciblée
# =====================
@main_bp.route("/admin/logs/delete", methods=["POST"])
@login_required
def admin_logs_delete():
    user = get_current_user()
    if not getattr(user, "is_admin", False):
        flash("⛔ Accès réservé à l'administrateur.", "danger")
        return redirect(url_for("main_bp.dashboard"))

    if request.form.get("delete_all"):
        deleted = AuditLog.query.delete()
        db.session.commit()
        flash("🧹 L’intégralité du journal d’audit a été supprimée.", "info")
        return redirect(request.form.get("next") or url_for("main_bp.admin_logs"))

    ids = []
    for raw in request.form.getlist("log_ids"):
        try:
            ids.append(int(raw))
        except (TypeError, ValueError):
            continue

    deleted = 0
    if ids:
        deleted = AuditLog.query.filter(AuditLog.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()

    if deleted:
        flash(f"🗑️ {deleted} entrée(s) supprimée(s) du journal.", "info")
    else:
        flash("Aucune entrée sélectionnée.", "warning")

    redirect_url = request.form.get("next") or url_for("main_bp.admin_logs")
    return redirect(redirect_url)


# =====================
# Timeline: ajout d'un commentaire
# =====================
@main_bp.route("/fiche/<int:fiche_id>/timeline/add", methods=["POST"])
@login_required
def add_timeline_comment(fiche_id):
    user = get_current_user()
    fiche = FicheImplique.query.get_or_404(fiche_id)
    # TODO: autorisations fines si besoin (mêmes règles que l'édition de fiche)
    raw_content = (request.form.get("comment") or "")
    content = re.sub(r"\r\n?", "\n", raw_content).strip()
    if not content:
        flash("Le commentaire est vide.", "warning")
        return redirect(url_for("main_bp.fiche_detail", fiche_id=fiche_id))
    if len(content) > TIMELINE_COMMENT_MAX_LENGTH:
        content = content[:TIMELINE_COMMENT_MAX_LENGTH].rstrip()
        flash(
            f"Le commentaire a été tronqué à {TIMELINE_COMMENT_MAX_LENGTH} caractères pour respecter la taille maximale.",
            "warning",
        )
    entry = TimelineEntry(fiche_id=fiche.id, user_id=user.id, content=content, kind="comment")
    db.session.add(entry)
    db.session.commit()
    log_action("timeline_add", "FicheImplique", fiche.id, extra=content[:200])
    flash("Commentaire ajouté.", "success")
    return redirect(url_for("main_bp.fiche_detail", id=fiche_id))

