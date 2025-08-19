from flask import render_template, request, redirect, url_for, session, flash, abort, jsonify, make_response
from datetime import datetime, timedelta
from flask_login import current_user
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
import io
from ..models import Utilisateur, Evenement, FicheImplique, Bagage, ShareLink, Ticket, Animal, utilisateur_evenement
from .. import db
from ..backup_utils import is_db_empty, backup_to_bytesio, wipe_db, bulk_restore


def register(bp):
    @bp.route("/evenement/new", methods=["GET", "POST"])
    @login_required
    def evenement_new():
        user = get_current_user()

        # 🔒 Restriction stricte à admin ou codep
        if not user.is_admin and user.role != "codep":
            flash("⛔ Vous n’avez pas l’autorisation de créer un évènement.", "danger")
            evenements = user.evenements  # on peut quand même lui afficher ceux qu’il voit
            return render_template("evenement_new.html", user=user, evenements=evenements)

        if request.method == "POST":
            nom_evt = request.form["nom_evt"]
            type_evt = request.form["type_evt"]
            adresse = request.form["adresse"]
            statut = request.form["statut"]

            # Génération du numéro d'évènement
            last_evt = Evenement.query.order_by(Evenement.id.desc()).first()
            next_id = last_evt.id + 1 if last_evt else 1
            numero_evt = str(next_id).zfill(8)

            # Création de l'évènement
            nouvel_evt = Evenement(
                numero=numero_evt,
                nom=nom_evt,
                type_evt=type_evt,
                adresse=adresse,
                statut=statut,
                createur_id=user.id,
                date_ouverture=datetime.utcnow()
            )

            db.session.add(nouvel_evt)
            db.session.commit()

            # Association du créateur à l'évènement
            if nouvel_evt not in user.evenements:
                user.evenements.append(nouvel_evt)
                db.session.commit()

            flash("✅ Évènement créé avec succès.", "success")
            return redirect(url_for("main_bp.dashboard", evenement_id=nouvel_evt.id))

        # 🔁 Méthode GET
        evenements = Evenement.query.all() if user.is_admin or user.role == "codep" else user.evenements
        return render_template("evenement_new.html", user=user, evenements=evenements)




    @bp.route("/evenement/<int:evenement_id>/dashboard")
    @login_required
    def dashboard(evenement_id):
        session["evenement_id"] = evenement_id
        user = get_current_user()

        evenement = Evenement.query.get(evenement_id)
        if not evenement or evenement not in user.evenements:
            flash("⛔ Vous n’avez pas accès à cet évènement.", "danger")
            return redirect(url_for("main_bp.evenement_new"))

        fiches = FicheImplique.query.filter_by(evenement_id=evenement.id).all()
        nb_present = FicheImplique.query.filter_by(evenement_id=evenement.id, statut="présent").count()
        nb_total = len(fiches)

        peut_modifier_statut = (
            user.is_admin or
            user.role == "codep" or
            evenement.createur_id == user.id or
            (user.role == "responsable" and user in evenement.utilisateurs)
        )

        return render_template(
            "dashboard.html",
            user=user,
            evenement=evenement,
            fiches=fiches,
            nb_present=nb_present,
            nb_total=nb_total,
            peut_modifier_statut=peut_modifier_statut,
            competence_colors=COMPETENCE_COLORS
        )








    # 🔁 Sélection d’un événement existant

    @bp.route("/evenement/select", methods=["POST"])
    @login_required
    def select_evenement():
        user = get_current_user()
        evt_id = request.form.get("evenement_id")

        if evt_id:
            session["evenement_id"] = int(evt_id)  # 🧠 on stocke dans la session
            return redirect(url_for("main_bp.dashboard", evenement_id=int(evt_id)))
        else:
            flash("Veuillez sélectionner un événement.", "warning")
            return redirect(url_for("main_bp.evenement_new"))





    # ➕ Création fiche impliqué (NOUVELLE VERSION)

    @bp.route("/evenement/<int:evenement_id>/update_statut", methods=["POST"])
    @login_required
    def update_evenement_statut(evenement_id):
        user = get_current_user()
        evenement = Evenement.query.get_or_404(evenement_id)

        if evenement not in user.evenements and not user.is_admin:
            flash("⛔ Accès refusé.", "danger")
            return redirect(url_for("main_bp.evenement_new"))

        new_statut = request.form.get("statut_evt")
        if new_statut:
            evenement.statut = new_statut
            db.session.commit()
            flash("✅ Statut de l’évènement mis à jour.", "success")

        return redirect(url_for("main_bp.dashboard", evenement_id=evenement.id))

    #############################################

    @bp.route("/evenement/<int:evenement_id>/fiches_json")
    @login_required
    def fiches_json(evenement_id):
        user = get_current_user()
        evt = Evenement.query.get_or_404(evenement_id)

        # 🔐 Sécurité : on vérifie l'accès à l'évènement
        if (evt not in user.evenements) and (not user.is_admin) and (user.role != "codep"):
            return jsonify({"error": "unauthorized"}), 403

        fiches = FicheImplique.query.filter_by(evenement_id=evenement_id).all()

        # Helper pour formatter les heures locales si dispos
        def fmt(dt):
            try:
                return dt.strftime('%d/%m/%Y %H:%M') if dt else "-"
            except Exception:
                return "-"

        fiches_data = []
        for fiche in fiches:
            # On suppose que les propriétés *_locale existent sur le modèle
            heure_arrivee_loc = getattr(fiche, "heure_arrivee_locale", None)
            heure_sortie_loc = getattr(fiche, "heure_sortie_locale", None)

            fiches_data.append({
                "id": fiche.id,
                "numero": fiche.numero,
                "nom": fiche.nom,
                "prenom": fiche.prenom,
                "statut": fiche.statut,
                "heure_arrivee": fmt(heure_arrivee_loc),
                "heure_sortie": fmt(heure_sortie_loc),
                "destination": fiche.destination or "",
                "difficultes": fiche.difficultes or "",
                "competences": fiche.competences or ""
            })

        # Event meta pour maj live (adresse + statut + date ouverture)
        date_ouv_loc = getattr(evt, "date_ouverture_locale", None)
        evenement_payload = {
            "id": evt.id,
            "nom": evt.nom,
            "adresse": evt.adresse or "",
            "statut": evt.statut or "",
            "date_ouverture": fmt(date_ouv_loc)
        }

        return jsonify({
            "fiches": fiches_data,
            "nb_present": sum(1 for f in fiches if f.statut == "présent"),
            "nb_total": len(fiches),
            "evenement": evenement_payload
        })


    #####################################################################

    COMPETENCE_COLORS = {
        "Médecin": "#e74c3c",
        "Infirmier": "#3498db",
        "Sapeur-pompier": "#e67e22",
        "SST": "#1abc9c",
        "Psychologue": "#9b59b6",
        "Bénévole": "#34495e",
        "Artisan": "#f39c12",
        "Interprète": "#2ecc71",
        "Logisticien": "#16a085",
        "Conducteur": "#d35400",
        "Agent sécurité": "#2c3e50",
        "Autre": "#7f8c8d"
    }



    #############################################"

    def _styled_table(data):
        table = Table(data, colWidths=[60*mm, 100*mm])
        table.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 11),
            ('BACKGROUND', (0,0), (-1,-1), colors.whitesmoke),
            ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.whitesmoke, colors.lightgrey]),
            ('TEXTCOLOR', (0,0), (-1,-1), colors.black),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('INNERGRID', (0,0), (-1,-1), 0.3, colors.grey),
            ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 6),
        ]))
        return table


    @bp.route('/evenement/<int:evenement_id>/edit', methods=['GET', 'POST'])
    @login_required
    def edit_evenement(evenement_id):
        user = get_current_user()
        evenement = Evenement.query.get_or_404(evenement_id)

        if not user.is_admin and user.role != "codep" and evenement.createur_id != user.id:
            flash("⛔ Accès interdit.", "danger")
            return redirect(url_for("main_bp.admin_evenements"))

        if request.method == "POST":
            evenement.nom = request.form["nom"]
            evenement.adresse = request.form["adresse"]
            evenement.type_evt = request.form["type"]
            evenement.statut = request.form["statut"]
            db.session.commit()
            flash("✅ Évènement mis à jour.", "success")
            return redirect(url_for("main_bp.admin_evenements"))

        return render_template("edit_evenement.html", evenement=evenement, user=user)

    #########################################


    @bp.route("/evenements/<int:evenement_id>/supprimer", methods=["POST"])
    @login_required
    def delete_evenement(evenement_id):
        user = get_current_user()  # ✅ au lieu de current_user
        evt = Evenement.query.get_or_404(evenement_id)

        # 🔐 Vérifie si l'utilisateur est admin OU le créateur (codep)
        if not (user.is_admin or user.role == "codep" or evt.createur_id == user.id):
            abort(403)

        # 🧹 Supprime les fiches impliquées
        FicheImplique.query.filter_by(evenement_id=evt.id).delete()

        # 🧹 Supprime les tickets (si tu en as)
        from .models import Ticket
        Ticket.query.filter_by(evenement_id=evt.id).delete()

        # 🗑 Supprime l'évènement
        db.session.delete(evt)
        db.session.commit()

        flash("✅ L’évènement et ses fiches ont été supprimés.", "success")
        return redirect(url_for("main_bp.evenement_new"))



    ###################################################



    @bp.route("/evenement/<int:evenement_id>/export/pdf")
    @login_required
    def export_evenement_fiches_pdf(evenement_id):
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        import io
        import pytz

        evenement = Evenement.query.get_or_404(evenement_id)
        fiches = FicheImplique.query.filter_by(evenement_id=evenement_id).all()

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
        elements = []

        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='CenterTitle', alignment=1, fontSize=18, spaceAfter=20))
        styles.add(ParagraphStyle(name='SubHeader', textColor=colors.orange, fontSize=14, spaceAfter=10))

        elements.append(Paragraph("Fiches Impliqués – Évènement", styles['CenterTitle']))
        elements.append(Paragraph("Informations sur l’évènement", styles['SubHeader']))

        # Date locale
        def convertir_heure_locale(dt_utc):
            if not dt_utc:
                return "Non renseignée"
            paris = pytz.timezone("Europe/Paris")
            return dt_utc.astimezone(paris).strftime("%d/%m/%Y %H:%M")

        infos_evt = [
            ["Nom", evenement.nom],
            ["Numéro", evenement.numero],
            ["Adresse", evenement.adresse],
            ["Statut", evenement.statut],
            ["Type", evenement.type_evt],
            ["Date d'ouverture", convertir_heure_locale(evenement.date_ouverture)]
        ]
        table_evt = Table(infos_evt, hAlign='LEFT', colWidths=[4*cm, 12*cm])
        table_evt.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        elements.append(table_evt)
        elements.append(Spacer(1, 20))

        elements.append(Paragraph("Liste des fiches impliquées", styles['SubHeader']))

        header = [
            "Nom", "Prénom", "Naissance", "Nationalité", "Statut",
            "Téléphone", "Adresse", "Compétences", "Destination", "Effets perso"
        ]
        data = [header]

        for f in fiches:
            row = [
                f.nom or "-",
                f.prenom or "-",
                f.date_naissance.strftime("%d/%m/%Y") if f.date_naissance else "-",
                f.nationalite or "-",
                f.statut or "-",
                f.telephone or "-",
                f.adresse or "-",
                f.competences or "-",
                f.destination or "-",
                f.effets_perso or "-",
            ]
            data.append(row)

        table_fiches = Table(data, repeatRows=1, hAlign='LEFT')
        table_fiches.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey)
        ]))
        elements.append(table_fiches)

        doc.build(elements)
        buffer.seek(0)

        return send_file(buffer, as_attachment=True, download_name=f"evenement_{evenement.numero}_fiches.pdf", mimetype='application/pdf')



    ###########################################



    @bp.route("/evenement/<int:evenement_id>/export/csv")
    @login_required
    def export_evenement_fiches_csv(evenement_id):
        # -> Désormais export XLSX stylé
        import io
        from datetime import datetime
        import pytz
        from flask import send_file, redirect, url_for, flash
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        user = get_current_user()
        evenement = Evenement.query.get_or_404(evenement_id)

        # Permissions : admin, codep, responsable rattaché
        if not (
            user.is_admin
            or user.role == "codep"
            or (user.role == "responsable" and user in evenement.utilisateurs)
        ):
            flash("⛔ Accès refusé pour l’export.", "danger")
            return redirect(url_for("main_bp.dashboard", evenement_id=evenement.id))

        # Fiches
        fiches = (
            FicheImplique.query
            .filter_by(evenement_id=evenement.id)
            .order_by(FicheImplique.id.asc())
            .all()
        )

        # Comptes
        nb_total = len(fiches)
        nb_present = sum(1 for f in fiches if (f.statut or "").lower() == "présent")
        nb_sorti = sum(1 for f in fiches if (f.statut or "").lower() == "sorti")

        # Timezone Paris
        paris = pytz.timezone("Europe/Paris")
        def to_paris_dt(dt):
            if not dt: return None
            try:
                return dt.astimezone(paris).replace(tzinfo=None)
            except Exception:
                return None

        # ====== Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Fiches Impliqués"

        # Couleurs
        BLEU = "002F6C"
        ORANGE = "F58220"
        GRIS_LIGNE = "E9EDF3"
        ZEBRA = "F8FAFF"

        # Styles
        th_font = Font(bold=True, color="FFFFFF")
        th_fill = PatternFill("solid", fgColor=BLEU)
        title_font = Font(bold=True, color="FFFFFF", size=16)
        banner_fill = PatternFill("solid", fgColor=ORANGE)
        key_cell = PatternFill("solid", fgColor="FFF3E6")
        val_cell = PatternFill("solid", fgColor="FFFFFFFF")
        txt_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        txt_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border_thin = Border(
            left=Side(style="thin", color=GRIS_LIGNE),
            right=Side(style="thin", color=GRIS_LIGNE),
            top=Side(style="thin", color=GRIS_LIGNE),
            bottom=Side(style="thin", color=GRIS_LIGNE),
        )

        # ====== En-tête événement
        # Titre bandeau
        headers = [
            "Numéro", "Code Sinus", "Nom", "Prénom", "Date de naissance", "Téléphone",
            "Adresse", "Statut", "Heure d’arrivée", "Heure de sortie", "Destination",
            "Moyen de transport", "Recherche personne", "N° recherche",
            "Personne à prévenir", "Tél. à prévenir", "Difficultés",
            "Compétences", "Bagages", "Autres informations",
        ]
        last_col = get_column_letter(len(headers))

        ws.merge_cells(f"A1:{last_col}1")
        c = ws["A1"]
        c.value = "📋 Export Fiches Impliqués — Protection Civile"
        c.font = title_font
        c.alignment = txt_left
        c.fill = banner_fill
        ws.row_dimensions[1].height = 26

        # Tableau d’infos évènement (2 colonnes: clé / valeur) sur 2 colonnes x 4 lignes (8 infos)
        evt_pairs = [
            ("Évènement", evenement.nom or ""),
            ("Numéro", evenement.numero or ""),
            ("Adresse", evenement.adresse or ""),
            ("Statut", evenement.statut or ""),
            ("Type", evenement.type_evt or ""),
            ("Ouverture", to_paris_dt(evenement.date_ouverture)),
            ("Présents", nb_present),
            ("Total / Sortis", f"{nb_total} / {nb_sorti}"),
        ]

        start_row = 3
        for idx, (k, v) in enumerate(evt_pairs):
            r = start_row + idx
            # clé
            ws[f"A{r}"].value = k
            ws[f"A{r}"].fill = key_cell
            ws[f"A{r}"].font = Font(bold=True, color=BLEU)
            ws[f"A{r}"].alignment = txt_left
            ws[f"A{r}"].border = border_thin
            # valeur (colonne B fusionnée jusqu’à D pour laisser de l'espace)
            ws.merge_cells(f"B{r}:D{r}")
            cell = ws[f"B{r}"]
            if isinstance(v, datetime):
                cell.value = v
                cell.number_format = "DD/MM/YYYY HH:MM"
            else:
                cell.value = v
            cell.fill = val_cell
            cell.alignment = txt_left
            cell.border = border_thin

        # Ligne vide
        table_start_row = start_row + len(evt_pairs) + 2

        # ====== En-têtes du tableau
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=table_start_row, column=col_idx, value=h)
            cell.font = th_font
            cell.fill = th_fill
            cell.alignment = txt_center
            cell.border = border_thin
        ws.freeze_panes = ws[f"A{table_start_row+1}"]  # fige titres
        ws.auto_filter.ref = f"A{table_start_row}:{last_col}{table_start_row}"

        # ====== Lignes
        for i, f in enumerate(fiches, start=1):
            r = table_start_row + i
            # Bagages
            try:
                bag_nums = [b.numero for b in (f.bagages or []) if b and b.numero]
                bagages_txt = ", ".join(sorted(bag_nums))
            except Exception:
                bagages_txt = ""

            # Dates/Heures (format Excel)
            d_naiss = f.date_naissance  # date ou None
            h_arr = to_paris_dt(getattr(f, "heure_arrivee", None))
            # si propriété *_locale dispo:
            if getattr(f, "heure_arrivee_locale", None):
                h_arr = f.heure_arrivee_locale.replace(tzinfo=None)
            h_sort = to_paris_dt(getattr(f, "heure_sortie", None))
            if getattr(f, "heure_sortie_locale", None):
                h_sort = f.heure_sortie_locale.replace(tzinfo=None)

            row_vals = [
                f.numero or "",
                getattr(f, "code_sinus", "") or "",
                f.nom or "",
                f.prenom or "",
                d_naiss,                # Excel date
                f.telephone or "",
                f.adresse or "",
                f.statut or "",
                h_arr,                  # Excel datetime
                h_sort,                 # Excel datetime
                f.destination or "",
                f.moyen_transport or "",
                f.recherche_personne or "",
                getattr(f, "numero_recherche", "") or "",
                f.personne_a_prevenir or "",
                f.tel_personne_a_prevenir or "",
                f.difficultes or "",
                f.competences or "",
                bagages_txt,
                f.autres_informations or "",
            ]

            for c_idx, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=r, column=c_idx, value=val)
                cell.alignment = txt_left
                cell.border = border_thin
                # formats
                if c_idx == 5 and isinstance(val, datetime):
                    cell.number_format = "DD/MM/YYYY"
                if c_idx in (9, 10) and isinstance(val, datetime):
                    cell.number_format = "DD/MM/YYYY HH:MM"
                # zébrage
                if i % 2 == 1:
                    cell.fill = PatternFill("solid", fgColor=ZEBRA)

        # ====== Largeurs de colonnes (preset + auto approx)
        preset_widths = {
            "A": 12,  # Numéro
            "B": 18,  # Code Sinus
            "C": 20,  # Nom
            "D": 18,  # Prénom
            "E": 14,  # Naissance
            "F": 16,  # Téléphone
            "G": 30,  # Adresse
            "H": 12,  # Statut
            "I": 18,  # Arrivée
            "J": 18,  # Sortie
            "K": 22,  # Destination
            "L": 18,  # Moyen
            "M": 24,  # Recherche personne
            "N": 18,  # N° recherche
            "O": 24,  # Personne à prévenir
            "P": 18,  # Tél à prévenir
            "Q": 28,  # Difficultés
            "R": 28,  # Compétences
            "S": 24,  # Bagages
            "T": 32,  # Autres informations
        }
        for col, w in preset_widths.items():
            ws.column_dimensions[col].width = w

        # ====== Export
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        filename = f"evenement_{evenement.numero or evenement.id}_fiches.xlsx"
        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )



    ##################################################################


    def can_manage_sharing(user):
        return user.is_admin or user.role in {"codep", "responsable"}

    # Vue autorité (avec panneau de gestion des liens, login requis)

    @bp.route("/evenement/<int:evenement_id>/autorite", methods=["GET"])
    @login_required
    def autorite_dashboard_manage(evenement_id):
        user = get_current_user()
        evt = Evenement.query.get_or_404(evenement_id)

        # accès : admin/codep/responsable ou créateur
        if not (user.is_admin or user.role in {"codep", "responsable"} or evt.createur_id == user.id):
            flash("⛔ Accès refusé.", "danger")
            return redirect(url_for("main_bp.dashboard", evenement_id=evenement_id))

        links = ShareLink.query.filter_by(evenement_id=evenement_id).order_by(ShareLink.created_at.desc()).all()
        return render_template("autorite_dashboard.html", user=user, evenement=evt, links=links, manage=True)

    # Création d’un lien (login requis)

    @bp.route("/evenement/<int:evenement_id>/share/create", methods=["POST"])
    @login_required
    def create_share_link(evenement_id):
        user = get_current_user()
        evt = Evenement.query.get_or_404(evenement_id)
        if not can_manage_sharing(user):
            abort(403)
        # durée optionnelle (en heures), vide = sans expiration
        hours = (request.form.get("duration_hours") or "").strip()
        expires_at = None
        if hours.isdigit():
            expires_at = datetime.utcnow() + timedelta(hours=int(hours))

        token = ShareLink.new_token()
        link = ShareLink(token=token, evenement_id=evt.id, created_by=user.id, expires_at=expires_at)
        db.session.add(link)
        db.session.commit()
        flash("🔗 Lien de partage créé.", "success")
        return redirect(url_for("main_bp.autorite_dashboard_manage", evenement_id=evt.id))

    # Révocation d’un lien (login requis)

    @bp.route("/evenement/<int:evenement_id>/autorite_json")
    def autorite_json(evenement_id):
        token = request.args.get("token")
        evt = Evenement.query.get_or_404(evenement_id)

        if token:
            link = ShareLink.query.filter_by(token=token, evenement_id=evenement_id).first()
            if not link or not link.is_active():
                # 403 si invalide; côté client on reste avec '—'
                return jsonify({"error":"forbidden"}), 403
        else:
            # chemin connecté (dashboard opérateur)
            if "user_id" not in session:
                return jsonify({"error":"unauthorized"}), 401
            user = get_current_user()
            if (evt not in user.evenements) and not (user.is_admin or user.role in {"codep","responsable"}):
                return jsonify({"error":"forbidden"}), 403

        fiches = FicheImplique.query.filter_by(evenement_id=evenement_id).all()
        nb_present = sum(1 for f in fiches if f.statut == "présent")
        nb_sorti   = sum(1 for f in fiches if f.statut == "sorti")
        nb_total   = len(fiches)

        return jsonify({
            "evenement": {
                "id": evt.id,
                "nom": evt.nom or "",
                "adresse": evt.adresse or "",
                "statut": evt.statut or "",
                "date_ouverture": (evt.date_ouverture_locale.strftime("%d/%m/%Y %H:%M")
                                   if getattr(evt, "date_ouverture_locale", None) else "")
            },
            "stats": {
                "nb_present": nb_present,
                "nb_sorti": nb_sorti,
                "nb_total": nb_total
            }
        })





    #######################



    def has_ticket_rights(user):
        return bool(
            user.is_admin or
            (user.role or "").lower() in {"codep", "responsable", "logisticien"}
        )


    # ===== TICKETS =====

    @bp.route("/evenement/<int:evenement_id>/tickets")
    @login_required
    def tickets_board(evenement_id):
        user = get_current_user()
        evenement = Evenement.query.get_or_404(evenement_id)
        users = evenement.utilisateurs  # ou ta logique de sélection

        can_manage = user.is_admin or user.role in ["codep", "responsable", "logisticien"]

        users_data = [
            {"id": u.id, "nom": u.nom, "role": u.role}
            for u in users
        ]

        return render_template(
            "tickets_board.html",
            evenement=evenement,
            users=users,
            users_data=users_data,  # ✅ on passe ici
            can_manage=can_manage,
            user=user
        )



    @bp.route("/evenement/<int:evenement_id>/tickets_json")
    @login_required
    def tickets_json(evenement_id):
        user = get_current_user()
        evt = Evenement.query.get_or_404(evenement_id)

        # Accès lecture : toute personne ayant accès à l'évènement ou admin
        if (evt not in user.evenements) and (not user.is_admin):
            return jsonify({"error": "forbidden"}), 403

        # Récup + tri (du plus récent au plus ancien)
        tickets = (
            Ticket.query
            .filter_by(evenement_id=evt.id)
            .order_by(Ticket.created_at.desc())
            .all()
        )

        return jsonify({
            "tickets": [t.to_dict() for t in tickets]
        })


